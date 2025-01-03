import random
from random_words import RandomWords
#pip install RandomWords
from DrissionPage import ChromiumPage,ChromiumOptions
from loguru import logger
import random
# 连接数据库
from flaskServer.config.connect import app
#数据库信息
from flaskServer.mode.env import Env
import time
#配置代理
from flaskServer.mode.proxy import Proxy
#创建浏览器
from flaskServer.services.chromes.worker import submit
#变量
from flaskServer.services.content import Content
#登录环境账号
from flaskServer.services.chromes.login import OKXChrome
from flaskServer.services.dto.account import getAccountById
from pprint import pprint
from flaskServer.config.connect import db
from flaskServer.mode.account import Account
from flaskServer.utils.crypt import aesCbcPbkdf2DecryptFromBase64
from flaskServer.services.chromes.login import tw2faV
from flaskServer.services.dto.env import updateAllStatus,getAllEnvs,getEnvsByGroup
from threading import Thread
from flaskServer.services.chromes.login import LoginDiscord
from flaskServer.utils.chrome import quitChrome
from flaskServer.utils.decorator import chrome_retry
from flaskServer.services.dto.task_record import updateTaskRecord, getTaskObject
from flaskServer.services.chromes.login import LoginTW
import openpyxl
import time
name = "Deek"
click_wallet_js = """
            const button = document.querySelector("body > w3m-modal").shadowRoot.querySelector("wui-flex > wui-card > w3m-router").shadowRoot.querySelector("div > w3m-connect-view").shadowRoot.querySelector("wui-flex > w3m-wallet-login-list").shadowRoot.querySelector("wui-flex > w3m-connect-injected-widget").shadowRoot.querySelector("wui-flex > wui-list-wallet:nth-child(1)").shadowRoot.querySelector("button > wui-text").shadowRoot.querySelector("slot");
            return button
            """
deek_network_js = """
            const button = document.querySelector("body > w3m-modal").shadowRoot.querySelector("wui-flex > wui-card > w3m-router").shadowRoot.querySelector("div > w3m-unsupported-chain-view").shadowRoot.querySelector("wui-flex > wui-flex:nth-child(2) > wui-list-network").shadowRoot.querySelector("button");            
            return button
            """


def getTab(chrome, env):
    tab = chrome.new_tab(url="https://www.deek.network/")
    chrome.wait(2, 3)

    if tab.ele('Join now'):
        tab.ele('Join now').click()
        chrome.wait(3, 6)

    if tab.ele('t:span@text():Login with X'):
        logger.info(f"{env.name}开始授权 X")
        tab.ele('t:span@text():Login with X').click()
        chrome.wait(30, 35)
        for _ in range(3):
            if tab.ele('t:div@text():please try again'):
                chrome.get_tab(url='https://api.x.com/').close()
                tab.refresh()
                chrome.wait(2)
                tab.ele('t:span@text():Login with X').click()
                chrome.wait(30, 35)

        try:
            tw_tab = chrome.get_tab(url="twitter")
            if tw_tab:
                    if "login" in tw_tab.url:
                        logger.info(f"{env.name}: 推特未登录,尝试重新登录")
                        tw: Account = getAccountById(env.tw_id)
                        if tw:
                            tw_tab.ele("@autocomplete=username").input(tw.name)
                            tw_tab.ele("@@type=button@@text()=Next").click()
                            tab.ele("@type=password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd))
                            tw_tab.ele("@@type=button@@text()=Log in").click()
                            fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                            if "login" in tab.url and len(fa2) > 10:
                                tw2faV(tab, fa2)
                            chrome.wait(25, 30)
                        else:
                                raise Exception(f"{env.name}: 没有导入TW的账号信息")
        except Exception as e:
            logger.info(f"y{env.name}: 推特登陆失败")
            return

        try:
            if chrome.get_tab(url='https://twitter.com/').s_ele("@@type=submit@@value=Send email"):
                logger.info(f"{env.name}   该环境推特需要邮箱验证，请前往验证")
                quitChrome(env, chrome)
        except Exception as e:
            pass

        try:
                if chrome.get_tab(url='https://twitter.com/').s_ele("@@type=submit@@value=Start"):
                    chrome.get_tab(url='https://twitter.com/').ele("@@type=submit@@value=Start").click()
                    chrome.wait(10, 15)
                    if chrome.get_tab(url='https://twitter.com/').s_ele("@@type=submit@@value=Send email"):
                        logger.info(f"{env.name}   该环境推特需要邮箱验证，请前往验证")
                        quitChrome(env, chrome)
        except Exception as e:
            pass

        try:
                if chrome.get_tab(url='https://twitter.com/').ele("@@type=submit@@value=Continue to X"):
                            chrome.get_tab(url='https://twitter.com/').ele("@@type=submit@@value=Continue to X").click()
                            chrome.wait(20, 25)
                if chrome.get_tab(url='https://twitter.com/').s_ele("@@type=submit@@value=Start"):
                            chrome.get_tab(url='https://twitter.com/').ele("@@type=submit@@value=Start").click()
                            chrome.wait(10, 15)
                if chrome.get_tab(url='https://twitter.com/').s_ele("@@type=submit@@value=Send email"):
                            logger.info(f"{env.name}   该环境推特需要邮箱验证，请前往验证")
                            quitChrome(env, chrome)
        except Exception as e:
            pass

        try:
            logger.info(f"{env.name}授权 X")
            tab.wait.ele_displayed(chrome.get_tab(url='https://api.x.com/').ele("@class=submit button selected"), timeout=120)
            chrome.get_tab(url='https://api.x.com/').ele("@class=submit button selected").click()
            logger.info(f"{env.name}授权 X 完成")
        except Exception as e:
            pass

        try:
            tab.wait.ele_displayed('t:button@text():Create', timeout=120)
            if tab.ele('t:button@text():Create'):
                logger.info(f"{env.name}创建账户")
                chrome.wait(2, 3)
                tab.ele('t:button@text():Create').click()
                chrome.wait(2, 3)
                for _ in range(3):
                    if tab.ele('t:button@text():Create'):
                        tab.refresh()
                        chrome.wait(2, 3)
                        tab.ele('t:button@text():Create').click()
                        chrome.wait(1, 2)

                tab.wait.load_start()
                tab.ele('@placeholder=Enter invite code').click().input('13VRQ3', clear=True)
                logger.info(f"{env.name}输入邀请码")
                chrome.wait(1, 2)
                tab.ele('t:button@text():Confirm').click()
                logger.info(f"{env.name}提交邀请码")
                chrome.wait(3, 6)

        except Exception as e:
            try:
                if tab.ele('t:button@text():Create'):
                    chrome.wait(2, 3)
                    tab.ele('t:button@text():Create').click()
                    logger.info(f"{env.name}创建账户")
                    chrome.wait(2, 3)
                    for _ in range(3):
                        if tab.ele('t:button@text():Create'):
                            tab.refresh()
                            chrome.wait(2, 3)
                            tab.ele('t:button@text():Create').click()
                            chrome.wait(1, 2)

                    tab.wait.load_start()
                    tab.ele('@placeholder=Enter invite code').click().input('13VRQ3', clear=True)
                    logger.info(f"{env.name}输入邀请码")
                    chrome.wait(1, 2)
                    tab.ele('t:button@text():Confirm').click()
                    logger.info(f"{env.name}提交邀请码")
                    chrome.wait(3, 6)
            except Exception as e:
                pass

        if tab.ele('t:button@text():Connect Wallet'):
            logger.info(f"{env.name}主页登录成功")

def getDeek(chrome, env):
    tab = chrome.new_tab(url='https://www.deek.network/')
    chrome.wait(2, 4)
    logger.info(f"{env.name}  登陆钱包")
    chrome.wait(4, 8)

    try:
        tab.run_js(deek_network_js).click()
    except Exception as e:
        pass

    logger.info(f"{env.name}    判断是否要重新连接钱包")
    if tab.ele("t:button@tx():Connect Wallet"):
        tab.ele("t:button@tx():Connect Wallet").click()

        chrome.wait(4, 8)
        logger.info(f"{env.name}    选择okx钱包")
        tab.run_js(click_wallet_js).click()

        chrome.wait(4, 8)
        logger.info(f"{env.name}    okx钱包确认连接")
        chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
        chrome.wait(12, 14)

        try:
            tab.run_js(deek_network_js).click()
            chrome.wait(4, 8)
            tab.run_js(click_wallet_js).click()
            chrome.wait(4, 8)
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
        except Exception as e:
            pass
    else:
        logger.info(f"{env.name}  钱包已连接，等待授权...")

    chrome.wait(4, 8)
    if chrome.get_tab(title="OKX Wallet"):
        logger.info(f"{env.name}    okx钱包授权")
        chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()


    logger.info(f"{env.name} 关注推特触发登录")
    tab.ele('Go').click()
    chrome.wait(15, 20)
    if chrome.get_tab(url='https://discord.com/'):
        chrome.get_tab(url='https://discord.com/').close()
        return
    chrome.get_tab(url='https://x.com/').ele("@type=button").click()
    chrome.wait(2, 4)
    if tab.ele('t:span@text():Follow'):
        tab.ele('t:span@text():Follow').click()

    if chrome.get_tab().ele('t:span@text():Log in'):
        chrome.get_tab().ele('t:span@text():Log in').click()

    chrome.wait(3, 6)

    if chrome.get_tab(url='https://x.com/'):
        logger.info(f"{env.name} 开始验证任务1")
        tab.ele('@class=btn-secondary-medium self-stretch max-w-[156px] xs:min-w-[216px] md:min-w-[134px] xl:min-w-[118px] 2xl:max-w-[113px]').click()
        chrome.get_tab(url='https://x.com/').close()

    chrome.wait(12, 15)
    tw_tab = chrome.get_tab(url="twitter")
    if tw_tab:
        if "login" in tw_tab.url:
            logger.info(f"{env.name}: 推特未登录,尝试重新登录")
            with app.app_context():
                tw: Account = Account.query.filter_by(id=env.tw_id).first()
                if tw:
                    tw_tab.ele("@autocomplete=username").input(tw.name)
                    tw_tab.ele("@@type=button@@text()=Next").click()
                    tw_tab.ele("@type=password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd))
                    tw_tab.ele("@@type=button@@text()=Log in").click()
                    fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                    if "login" in tab.url and len(fa2) > 10:
                        tw2faV(tab, fa2)
                    chrome.wait(15, 20)
                    tw_tab.ele("Authorize app").click()
                    logger.info(f"{env.name}: 推特授权成功")
                else:
                    raise Exception(f"{env.name}: 没有导入TW的账号信息")
        else:
            tw_tab.ele("Authorize app").click()
            chrome.wait(4, 8)

    var = 1
    max_attempts = 12
    attempts = 0

    while var == 1 and attempts < max_attempts:
        logger.info(f"{env.name}  关注推特任务")
        if tab.ele('@class=btn-primary-medium self-stretch max-w-[156px] xs:min-w-[216px] md:min-w-[134px] xl:min-w-[118px] 2xl:max-w-[113px]'):
            tab.ele('@class=btn-primary-medium self-stretch max-w-[156px] xs:min-w-[216px] md:min-w-[134px] xl:min-w-[118px] 2xl:max-w-[113px]').click()
            chrome.wait(4, 8)
            if chrome.get_tab(url='https://discord.com/'):
                chrome.get_tab(url='https://discord.com/').close()
                return
        else:
            logger.info(f"{env.name} 任务已全部完成")

        if chrome.get_tab(url='https://x.com/'):

            chrome.wait(6, 8)
            if chrome.get_tab(url='https://x.com/'):
                chrome.get_tab(url='https://x.com/').ele("@type=button").click()
                chrome.wait(2, 4)
                if tab.ele('t:span@text():Follow'):
                    tab.ele('t:span@text():Follow').click()
                    chrome.wait(2, 4)

                if chrome.get_tab(url='https://x.com/'):
                    tab.ele('t:button@text():Verify').click()
                    logger.info(f"{env.name}  验证")
                    chrome.wait(5, 7)
                else:
                    pass

                if tab.ele('Verification Failed. Mostly due to twitter being delayed, please try again after 30 seconds'):
                    logger.info(f"{env.name}  关注推特任务出错")
                    if chrome.get_tab().ele('t:span@text():Log in'):
                        logger.info(f"{env.name}  推特登录失败，退出任务")
                        var = 0
                        return
                    tab.refresh()
                    if chrome.get_tab(title="OKX Wallet"):
                        logger.info(f"{env.name}  okx钱包授权")
                        chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()


                    chrome.get_tab(url='https://x.com/').ele("@type=button").click()
                    chrome.wait(10, 12)
                    tab.ele('t:button@text():Verify').click()
                    logger.info(f"{env.name}  验证")
                else:
                    logger.info(f"{env.name}  关注推特任务成功")
                    chrome.get_tab(url='https://x.com/').close()
        else:
            chrome.get_tab(url='https://discord.com/').close()
            tab.ele('@class=btn-primary-medium self-stretch max-w-[156px] xs:min-w-[216px] md:min-w-[134px] xl:min-w-[118px] 2xl:max-w-[113px]').next().click()
        attempts += 1

    return

def dailyTask(chrome, env):

        try:
            tab = chrome.new_tab(url='https://www.deek.network/')
            chrome.wait(3, 6)
            if chrome.get_tab(title="OKX Wallet"):
                logger.info(f"{env.name}  okx钱包授权")
                chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
                chrome.wait(10, 12)
            tab.ele('@class=btn-primary-medium self-stretch max-w-[156px] xs:min-w-[216px] md:min-w-[134px] xl:min-w-[118px] 2xl:max-w-[113px]', index=2).click(by_js=None)
            chrome.wait(10, 12)

            if chrome.get_tab(url='https://x.com/').ele('@data-testid=tweetButton'):
                logger.info(f"{env.name}  每日任务一")
                chrome.get_tab(url='https://x.com/').ele('@data-testid=tweetButton').click(by_js=None)
                chrome.wait(2, 3)
                chrome.get_tab(url='https://x.com/').close()
                chrome.wait(10, 12)
            tab.ele('@class=btn-primary-medium self-stretch max-w-[156px] xs:min-w-[216px] md:min-w-[134px] xl:min-w-[118px] 2xl:max-w-[113px]', index=3).click(by_js=None)
            chrome.wait(10, 12)
            chrome.get_tab(url='https://x.com/').close()
        except Exception as e:
            pass

        try:

            tab = chrome.new_tab(url='https://x.com/settings/profile')
            logger.info(f"{env.name}  每日任务二")
            chrome.wait(3, 5)
            tab.ele('@name=displayName').input(' DEEK')
            chrome.wait(2, 3)
            tab.ele('@data-testid=Profile_Save_Button').click(by_js=None)
            chrome.wait(3, 6)
            chrome.get_tab(url='https://x.com/').close()
            chrome.wait(10, 12)

        except Exception as e:
            pass

        tab = chrome.new_tab(url='https://www.deek.network/')
        if chrome.get_tab(title="OKX Wallet"):
            logger.info(f"{env.name}  okx钱包授权")
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
            chrome.wait(10, 12)

        logger.info(f"{env.name}  验证")
        try:
            tab.ele('t:button@text():Verify', index=2).click()
            chrome.wait(10, 12)
            tab.ele('t:button@text():Verify', index=2).click()
            chrome.wait(10, 12)
            logger.info(f"{env.name}  每日任务完成")
        except Exception as e:
            logger.info(f"{env.name}  每日任务完成")

        return

def deekCount(chrome, env):

    tab = chrome.new_tab(url='https://www.deek.network/')
    taskData = getTaskObject(env, name)
    env_name = env.name
    chrome.wait(2, 4)
    logger.info(f"{env.name}  登陆钱包")
    chrome.wait(4, 8)

    try:
        tab.run_js(deek_network_js).click()
    except Exception as e:
        pass

    logger.info(f"{env.name}    判断是否要重新连接钱包")
    if tab.ele("t:button@tx():Connect Wallet"):
        tab.ele("t:button@tx():Connect Wallet").click()

        chrome.wait(4, 8)
        logger.info(f"{env.name}    选择okx钱包")
        tab.run_js(click_wallet_js).click()

        chrome.wait(4, 8)
        logger.info(f"{env.name}    okx钱包确认连接")
        chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
        chrome.wait(12, 14)

        try:
            tab.run_js(deek_network_js).click()
            chrome.wait(4, 8)
            tab.run_js(click_wallet_js).click()
            chrome.wait(4, 8)
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
        except Exception as e:
            pass
    else:
        logger.info(f"{env.name}  钱包已连接，等待授权...")

    chrome.wait(4, 8)
    if chrome.get_tab(title="OKX Wallet"):
        logger.info(f"{env.name}    okx钱包授权")
        chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
        chrome.wait(4, 8)

    logger.info(f"{env.name}  开始数据统计")

    points = tab.ele('@class=text-15-s60-l60-w700 font-degular-display capitalize not-italic text-content-accent1').text
    top = tab.ele('@class=font-sf-pro-display text-5-s20-l30-w700 not-italic text-content-primary').text
    text = str(top)
    number = text.split()[2]

        # 使用原始字符串方式指定文件路径
    current_time = time.strftime("%m-%d")
    file_path = r'C:\Users\Public\Documents\deek_{}.xlsx'.format(current_time)

        # 打开已存在的 Excel 文件（arch.xlsx）
    try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = 'points'
            ws['C1'] = 'number'
    except FileNotFoundError:
            # 如果文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = 'points'
            ws['C1'] = 'number'
            wb.save(file_path)
        # 找到下一行位置（避免覆盖）
    next_row = ws.max_row + 1

    env_name_exists = False
    for row in range(2, ws.max_row + 1):  # 从第二行开始遍历（跳过表头）
        if ws[f'A{row}'].value == env_name:
                # 如果找到相同的 env_name，更新该行的 xp 和 level
                ws[f'B{row}'] = points
                ws[f'C{row}'] = number
                env_name_exists = True
                break
    if not env_name_exists:
            # 如果没有找到相同的 env_name，追加新行
            next_row = ws.max_row + 1
            ws[f'A{next_row}'] = env_name
            ws[f'B{next_row}'] = points
            ws[f'C{next_row}'] = number

        # 保存文件（不会覆盖，直接追加）
    wb.save(file_path)

    taskData.Points = points
    taskData.top = number
    updateTaskRecord(env.name, name, taskData, 1)
    tab.close()

    return

def deek(env):
    with app.app_context():
        try:
            chrome: ChromiumPage = OKXChrome(env)
            getTab(chrome, env)
            getDeek(chrome, env)
            dailyTask(chrome, env)
            deekCount(chrome, env)
            logger.info(f"{env.name}环境：任务执行完毕，关闭环境")
        except Exception as e:
            logger.error(f"{env.name} 执行：{e}")
            return ("失败", e)
        finally:
            quitChrome(env, chrome)


