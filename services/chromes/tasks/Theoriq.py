import openpyxl
from flaskServer.mode.account import Account
from flaskServer.services.chromes.login import OKXChrome, tw2faV
from DrissionPage import ChromiumPage
from loguru import logger
# 连接数据库
from flaskServer.config.connect import app
#登录环境账号
from flaskServer.services.chromes.login import OKXChrome
from flaskServer.services.dto.account import updateAccountStatus
from flaskServer.services.dto.task_record import updateTaskRecord, getTaskObject
import time
from flaskServer.utils.chrome import quitChrome, get_Custome_Tab
from flaskServer.utils.crypt import aesCbcPbkdf2DecryptFromBase64
import requests

name = 'Theoriq'
Theoriq_url = 'https://quests.theoriq.ai?r=NpRqz3aq'

okx_js = '''
let button =
document.querySelector("#privy-modal-content > div > div.sc-bbQqnZ.cvJuSO > div.sc-hbtGpV.kQPurx > button:nth-child(1)");
button.click();
'''
def LoginDiscord(chrome:ChromiumPage,env):
    updateAccountStatus(env.discord_id, 0, "重置了Discord登录状态")
    tab = chrome.new_tab(url="https://discord.com/app")
    if tab.s_ele("Please log in again"):
        tab.ele("@class=button_dd4f85 lookFilled_dd4f85 colorPrimary_dd4f85 sizeMedium_dd4f85 grow_dd4f85").click()
    if "login" in tab.url:
        logger.info(f"{env.name} 开始登录 Discord 账号")
        with app.app_context():
            discord:Account = Account.query.filter_by(id=env.discord_id).first()
            if discord:
                tab.ele("@name=email").input(discord.name)
                tab.ele("@name=password").input(aesCbcPbkdf2DecryptFromBase64(discord.pwd))
                tab.ele("@type=submit").click()
                fa2 = aesCbcPbkdf2DecryptFromBase64(discord.fa2)
                if "login" in tab.url and len(fa2) > 10:
                    res = requests.get(fa2)
                    if res.ok:
                        code = res.json().get("data").get("otp")
                        tab.ele("@autocomplete=one-time-code").input(code)
                        tab.ele("@type=submit").click()
            else:
                updateAccountStatus(env.discord_id, 1, "没有导入DISCORD 的账号信息")
                raise Exception(f"{env.name}: 没有导入DISCORD 账号信息")
    if "channels" in tab.url or ".com/app" in tab.url:
        updateAccountStatus(env.discord_id, 2)
        logger.info(f"{env.name}登录Discord成功！")
    return get_Custome_Tab(tab)


def getSigninTW(chrome, env):
    try:
        tab = chrome.new_tab(url='https://x.com/')
        chrome.wait(8, 10)
        logger.info('开始判断')

        # 登录处理逻辑
        if any(tab.s_ele(selector) for selector in ["Sign in", "Log in", "Retry", "Refish"]):
            logger.info(f"{env.name}: 推特未登录，触发登录推特")
            if tab.s_ele('Sign in'):
                tab.ele('Sign in').click()
            elif tab.s_ele('Log in'):
                tab.ele('Log in').click()
            elif tab.s_ele('Retry'):
                logger.info("Refreshing the page due to Retry")
                tab.refresh(ignore_cache=True)

            time.sleep(10)  # 等待页面加载

            with app.app_context():
                tw: Account = Account.query.filter_by(id=env.tw_id).first()
                if not tw:
                    raise Exception(f"{env.name}: 没有导入TW的账号信息")

                # 输入用户名和密码
                tab.ele("@autocomplete=username").input(tw.name, clear=True)
                tab.ele("@@type=button@@text()=Next").click()
                tab.ele("@type=password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd), clear=True)
                tab.ele("@@type=button@@text()=Log in").click()

                # 二次验证
                fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                if "login" in tab.url and len(fa2) > 10:
                    tw2faV(tab, fa2)

                tab.ele('@type=button').click()
                chrome.wait(2)
                logger.info(f'{env.name}: 登录完成')
                chrome.close_tabs()

        # 处理人工验证和其他状态
        elif tab.s_ele('@class=Button EdgeButton EdgeButton--primary') or tab.s_ele('@value=Start') or tab.s_ele('@value=Continue to X'):
            if tab.s_ele('@class=Button EdgeButton EdgeButton--primary'):
                logger.info(f'{env.name}:需要人工验证twitter，是Send email')
                time.sleep(1)
                return
            elif tab.s_ele('@value=Start'):
                tab.ele('@value=Start').click()
                time.sleep(15)
                if tab.s_ele('@value=Continue to X'):
                    tab.ele('@value=Continue to X').click()
                elif tab.s_ele('@value=Send email'):
                    logger.info(f'{env.name}:需要人工验证twitter,是Send email')
                    time.sleep(1)
                    return
                else:
                    time.sleep(10)
                    tab.ele('@value=Continue to X').click()
            elif tab.s_ele('@value=Continue to X'):
                tab.ele('@value=Continue to X').click()
        elif tab.s_ele('Your account is suspended'):
            print(f'{env.name}:此号被封')
            return False
    except Exception as e:
        logger.error(f"处理过程中出现错误: {e}")
        return


def exe_okx(chrome,env):
    try:
        if chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2):
            chrome.wait(3, 4)
            chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2).click()
            chrome.wait(3, 4)
            try:
                chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2).click()
            except Exception as e:
                print('不需要二次确认')
    except Exception as e:
        print(f'{env.name}取的ele不对或者不需要连接')
def getDiscord(chrome,env):
    try:
        print('开始discord认证')
        if chrome.get_tab(title='Discord | Authorize access to your account'):
            print('Discord | Authorize access to your account进入')
            chrome.get_tab(title='Discord | Authorize access to your account').ele("@type=button", index=2).click()
            logger.info(f"{env.name}: 登录discord完成----------------------------------")
            time.sleep(10)
        elif chrome.get_tab(title='Discord | 授权访问您的账号'):
            print('Discord | 授权访问您的账号进入')
            chrome.get_tab(title='Discord | 授权访问您的账号').ele("@type=button", index=2).click()
            logger.info(f"{env.name}: 登录discord完成----------------------------------")
            time.sleep(10)
        elif chrome.get_tab(title='Discord'):
            print('Discord 进入')
            if chrome.get_tab(title='Discord').s_ele('@class=button_dd4f85 lookFilled_dd4f85 colorPrimary_dd4f85 sizeMedium_dd4f85 grow_dd4f85'):
                logger.info(f'{env.name}的discord需要重新登录')
                chrome.get_tab(title='Discord').ele('@class=button_dd4f85 lookFilled_dd4f85 colorPrimary_dd4f85 sizeMedium_dd4f85 grow_dd4f85').click()
                time.sleep(6)
                LoginDiscord(chrome, env)
                time.sleep(6)
                chrome.close_tabs()
            elif chrome.get_tab(title='Discord').ele('Welcome back!'):
                logger.info(f"{env.name}的Discord未登录，尝试重新登录")
                time.sleep(6)
                LoginDiscord(chrome, env)
                time.sleep(6)
                chrome.close_tabs()
            if chrome.get_tab(title='Discord').ele("@type=button", index=2):
                chrome.get_tab(title='Discord').ele("@type=button", index=2).click()
                logger.info(f"{env.name}: 登录discord完成----------------------------------")
                time.sleep(10)
        else:
            logger.info(f'{env.name}:还有其他的语言需要加判断')
    except Exception as e:
        logger.error(e)


def getTab(chrome,env):
    tab = chrome.new_tab(url=Theoriq_url)
    tab.set.window.max()
    chrome.wait(5,10)
    try:
        num = 1
        while num < 5:
            if tab.s_ele('CONNECT WALLET'):
                print('点击次数：', num)
                tab.ele('CONNECT WALLET').click()
                print("CONNECT WALLET 点击完成")
                chrome.wait(2, 3)
                tab.ele('CONNECT METAMASK').click()
                print('CONNECT METAMASK点击完成')
                chrome.wait(2, 3)
                exe_okx(chrome, env)
                chrome.wait(5, 10)
                num += 1
            else:
                print('钱包请确认完成可以进入内场')
                break
    except Exception as e:
        logger.info(e)
    chrome.wait(2, 5)
    try:
        if tab.s_ele('@class=px-6 flex justify-between items-center gap-4 p-2 bg-[#181818] text-white rounded', index=2):
            print('判断成功开始点击')
            tab.ele('@class=px-6 flex justify-between items-center gap-4 p-2 bg-[#181818] text-white rounded', index=2).click()
            print('点击完成')
            chrome.wait(2, 3)
            exe_okx(chrome, env)
    except Exception as e:
        logger.error(e)

    chrome.wait(5, 10)
    try:
        print('开始discord确认')
        if tab.s_ele('Connect Discord'):
            tab.ele('Connect Discord').click()
            print('点击完成')
            chrome.wait(10, 20)
            print('title', chrome.get_tab(url='https://discord.com/').title)
            getDiscord(chrome, env)
            chrome.wait(2, 3)
    except Exception as e:
        logger.error(e)
    try:
        if tab.s_ele('@class=absolute right-4 top-4'):
            tab.ele('@class=absolute right-4 top-4').click()
    except Exception as e:
        logger.error(e)

    print('登录完成开始做任务')

def getSocialTasks(chrome,env):
    try:
        tw = getSigninTW(chrome, env)
        if tw == False:
            print('此号被封，无法进行下面任务了')
            quitChrome(env, chrome)
        else:
            pass
    except Exception as e:
        logger.error(e)
    tab = chrome.new_tab(url=Theoriq_url)
    chrome.wait(5, 10)
    try:
        if tab.s_ele('@class=px-3 pt-1 text-sm/[22px] font-medium text-black cursor-pointer text-center', index=1):
            tab.ele('@class=px-3 pt-1 text-sm/[22px] font-medium text-black cursor-pointer text-center',index=1).click()
            chrome.wait(5, 10)
            chrome.get_tab(url='https://api.x.com/').ele('@value=Authorize app').click()
            chrome.wait(5,13)
    except Exception as e:
        logger.error(e)
    try:
        tab.ele('t:button@tx():FOLLOW').click()
        chrome.wait(2, 3)
        chrome.get_tab(url='https://x.com/').ele("t:span@text():Follow").click()
        print('点击完成')
        chrome.wait(5, 13)
    except Exception as e:
        logger.info(e)



def getAgenttasks(chrome, env):
    tab = chrome.new_tab(url='https://infinity.theoriq.ai/login')
    try:
        chrome.wait(5, 10)
        tab.ele('Launch Infinity Studio').click()
        print('点击完成')
    except Exception as e:
        logger.info(e)

    chrome.wait(5, 10)
    try:
        print('Bondex Agent')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Bondex Agent')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=1).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)

    chrome.wait(2, 5)

    try:
        print('Filecoin Documentation Agent')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Filecoin Documentation Agent')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=1).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)

    chrome.wait(5, 8)

    try:
        print('Eigenlayer Discord Agent')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Eigenlayer Discord Agent')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=2).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)

    chrome.wait(5, 8)

    try:
        print('Eigenlayer Twitter Agent')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Eigenlayer Twitter Agent')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=2).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)

    chrome.wait(5, 8)
    try:
        print('Eigenlayer Documentation Agent')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Eigenlayer Documentation Agent')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=1).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)

    chrome.wait(5, 8)
    try:
        print('Eigenlayer Collective')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Eigenlayer Collective')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=2).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)
    chrome.wait(5, 8)

    chrome.wait(5, 8)
    try:
        print('Bondex Collective')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Bondex Collective')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=2).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)
    chrome.wait(5, 8)

    try:
        print('Bondex Collective')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Bondex Collective')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=2).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)
    chrome.wait(5, 8)
    try:
        print('Crypto Game')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('Crypto Game')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=2).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)

    chrome.wait(5, 8)
    try:
        print('DAO Agent')
        tab.ele('@class=size-6 text-green-500').click()
        chrome.wait(2, 3)
        tab.ele('@class=appearance-none w-full h-full bg-[transparent] border-none p-0 m-0 outline-none focus:outline-none').input('DAO Agent')
        chrome.wait(2, 3)
        tab.ele('@class=h-full px-1 lg:px-5').click()
        chrome.wait(2, 3)
        tab.ele('Create Session').click()
        chrome.wait(2, 3)
        tab.ele('@data-testid=suggested-question', index=2).click()
        chrome.wait(2, 3)
        tab.ele('@class=flex h-11 w-12 items-center justify-center rounded-full transition-colors duration-300 bg-green-500 hover:bg-green-300 ').click()
    except Exception as e:
        logger.info(e)

    return


def getCount(chrome, env):
    tab = chrome.new_tab(url='https://quests.theoriq.ai/leaderboard')
    try:
        taskData = getTaskObject(env, name)
        chrome.wait(10, 20)
        Your_rank = tab.ele('@class=font-bold text-3xl text-white', index=1).text
        print('Your_rank', Your_rank)
        Your_xp = tab.ele('@class=font-bold text-3xl text-white', index=2).text
        print('Your_xp', Your_xp)
        Completed_Quests = tab.ele('@class=font-bold text-3xl text-white', index=3).text
        print('Completed_Quests', Completed_Quests)

        taskData.Your_rank = Your_rank
        taskData.Your_xp = Your_xp
        taskData.Completed_Quests = Completed_Quests
        updateTaskRecord(env.name, name, taskData, 1)


        current_time = time.strftime("%m-%d")
        file_path = r'C:\Users\Public\Documents\theoriq_{}.xlsx'.format(current_time)

        # 打开已存在的 Excel 文件（arch.xlsx）
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = 'Your_rank'
            ws['C1'] = 'Your_xp'
            ws['D1'] = 'Completed_Quests'
        except FileNotFoundError:
            # 如果文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = 'Your_rank'
            ws['C1'] = 'Your_xp'
            ws['D1'] = 'Completed_Quests'
            wb.save(file_path)
        # 找到下一行位置（避免覆盖）
        next_row = ws.max_row + 1

        env_name_exists = False
        env_name = env.name
        for row in range(2, ws.max_row + 1):  # 从第二行开始遍历（跳过表头）
            if ws[f'A{row}'].value == env_name:
                # 如果找到相同的 env_name，更新该行的 xp 和 level
                ws[f'B{row}'] = Your_rank
                ws[f'C{row}'] = Your_xp
                ws[f'D{row}'] = Completed_Quests
                env_name_exists = True
                break
        if not env_name_exists:
            # 如果没有找到相同的 env_name，追加新行
            next_row = ws.max_row + 1
            ws[f'A{next_row}'] = env_name
            ws[f'B{next_row}'] = Your_rank
            ws[f'C{next_row}'] = Your_xp
            ws[f'D{next_row}'] = Completed_Quests
        # 保存文件（不会覆盖，直接追加）
        wb.save(file_path)

    except Exception as e:
        logger.error(e)



def theoriq(env):
    with app.app_context():
        try:
            chrome: ChromiumPage = OKXChrome(env)
            getTab(chrome, env)
            # getSocialTasks(chrome, env)
            getAgenttasks(chrome, env)
            getCount(chrome, env)
            logger.info(f"{env.name}环境：任务执行完毕，关闭环境")
        except Exception as e:
            logger.error(f"{env.name} 执行：{e}")
            return ("失败", e)
        finally:
            quitChrome(env, chrome)
