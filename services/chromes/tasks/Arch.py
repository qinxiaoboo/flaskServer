import random
from random_words import RandomWords
#pip install RandomWords
from DrissionPage import ChromiumPage,ChromiumOptions
from loguru import logger
import random
# 连接数据库
from flaskServer.config.connect import app
#数据库信息
from flaskServer.mode.env import Env
import time
#配置代理
from flaskServer.mode.proxy import Proxy
#创建浏览器
from flaskServer.services.chromes.worker import submit
#变量
from flaskServer.services.content import Content
#登录环境账号
from flaskServer.services.chromes.login import OKXChrome
from flaskServer.services.dto.account import getAccountById
from pprint import pprint
from flaskServer.config.connect import db
from flaskServer.mode.account import Account
from flaskServer.utils.crypt import aesCbcPbkdf2DecryptFromBase64
from flaskServer.services.chromes.login import tw2faV
from flaskServer.services.dto.env import updateAllStatus,getAllEnvs,getEnvsByGroup
from threading import Thread
from flaskServer.services.chromes.login import LoginDiscord
from flaskServer.utils.chrome import quitChrome
from flaskServer.utils.decorator import chrome_retry
from flaskServer.services.dto.task_record import updateTaskRecord, getTaskObject
from flaskServer.services.chromes.login import LoginTW
name = "Arch"
from flaskServer.utils.chrome import wait_captcha_page

def getTab(chrome, env):
    tab = chrome.new_tab(url="https://dashboard.arch.network?referralCode=f9c6ab90-03a4-4724-9cbd-080a192f74d2")
    rw = RandomWords()
    tab.set.window.max()
    chrome.wait(2, 3)

    if tab.ele('t:button@text():Connect Wallet'):
        logger.info(f"{env.name}   链接钱包选择OKX")
        tab.ele('t:button@text():Connect Wallet').click()
        chrome.wait(3, 6)
        tab.ele('OKX').click()
        chrome.wait(10, 15)

        if chrome.get_tab(title="OKX Wallet"):
            logger.info(f"{env.name}   OKX钱包授权")
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
            chrome.wait(10, 15)

    for _ in range(2):
        if tab.ele('t:button@text():Sign'):
            logger.info(f"{env.name}   登录")
            tab.ele('t:button@text():Sign').click()
            chrome.wait(5, 10)
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
            chrome.wait(10, 15)

    logger.info(f"{env.name}   进入任务页面")
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()

    tab.ele('t:span@text():START MISSIONS').click()
    chrome.wait(2, 3)
    tab.ele('t:div@text():WEEKLY MISSIONS').click()
    chrome.wait(2, 3)

    if tab.ele('t:button@text():Start'):
        logger.info(f"{env.name}   推特授权成功")
    else:
        logger.info(f"{env.name}   推特授权")
        try:
            tab.ele('t:button@text():Authorize').click()
        except Exception as e:
            pass
        chrome.wait(15, 20)

        try:
            tw_tab = chrome.get_tab(url="twitter")
            if tw_tab:
                if "login" in tw_tab.url:
                    logger.info(f"{env.name}: 推特未登录,尝试重新登录")
                    with app.app_context():
                        tw: Account = Account.query.filter_by(id=env.tw_id).first()
                        if tw:
                            tw_tab.ele("@autocomplete=username").input(tw.name, clear=True)
                            tw_tab.ele("@@type=button@@text()=Next").click()
                            tab.ele("t:span@text():Password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd), clear=True)
                            tw_tab.ele("@@type=button@@text()=Log in").click()
                            fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                            if "login" in tab.url and len(fa2) > 10:
                                tw2faV(tab, fa2)
                            chrome.wait(25, 30)
                        else:
                            raise Exception(f"{env.name}: 没有导入TW的账号信息")

            if chrome.get_tab(url='https://x.com/').ele("t:span@text():Log in"):
                chrome.get_tab(url='https://x.com/').ele("t:span@text():Log in").click()
                logger.info(f"{env.name}: 推特未登录,尝试重新登录")
                with app.app_context():
                    tw: Account = Account.query.filter_by(id=env.tw_id).first()
                    if tw:
                        tw_tab.ele("@autocomplete=username").input(tw.name, clear=True)
                        tw_tab.ele("@@type=button@@text()=Next").click()
                        tab.ele("t:span@text():Password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd), clear=True)
                        tw_tab.ele("@@type=button@@text()=Log in").click()
                        fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                        if "login" in tab.url and len(fa2) > 10:
                            tw2faV(tab, fa2)
                        chrome.wait(25, 30)
                    else:
                        raise Exception(f"{env.name}: 没有导入TW的账号信息")
        except Exception as e:
            pass

        try:
            if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Send email"):
                logger.info(f"{env.name}   该环境推特需要邮箱验证，请前往验证")
                quitChrome(env, chrome)
        except Exception as e:
            pass

        try:
            if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Start"):
                    chrome.get_tab(url='https://x.com/').ele("@@type=submit@@value=Start").click()
                    chrome.wait(10, 15)
                    if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Send email"):
                        logger.info(f"{env.name}   该环境推特需要邮箱验证，请前往验证")
                        quitChrome(env, chrome)
                    chrome.wait(25, 30)
        except Exception as e:
            pass

        try:
            if chrome.get_tab(url='https://x.com/').ele("@@type=submit@@value=Continue to X"):
                    chrome.get_tab(url='https://x.com/').ele("@@type=submit@@value=Continue to X").click()
                    chrome.wait(20, 25)
            if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Start"):
                chrome.get_tab(url='https://x.com/').ele("@@type=submit@@value=Start").click()
                chrome.wait(10, 15)
            if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Send email"):
                logger.info(f"{env.name}   该环境推特需要邮箱验证，请前往验证")
                quitChrome(env, chrome)
        except Exception as e:
            pass

        try:
            if chrome.get_tab(url='https://twitter.com/').ele("@data-testid=OAuth_Consent_Button"):
                    chrome.get_tab(url='https://twitter.com/').ele("@data-testid=OAuth_Consent_Button").click()
                    logger.info(f"{env.name}   推特授权成功")
                    chrome.wait(15, 20)
        except Exception as e:
                # tab = chrome.new_tab(url="https://dashboard.arch.network/missions")
                tab.ele('t:div@text():WEEKLY MISSIONS').click()
                chrome.wait(2, 4)

                if tab.ele('t:button@text():Start'):
                    logger.info(f"{env.name}   推特授权成功")
                else:
                    try:
                        tab.ele('t:button@text():Authorize').click()
                        chrome.wait(15, 20)
                        chrome.get_tab(url='https://twitter.com/').ele("@data-testid=OAuth_Consent_Button").click()
                        logger.info(f"{env.name}   推特授权成功")
                    except Exception as e:
                        logger.info(f"{env.name}   推特授权失败")
                        quitChrome(env, chrome)
    return

def missions(chrome, env):
    tab = chrome.new_tab(url="https://dashboard.arch.network?referralCode=f9c6ab90-03a4-4724-9cbd-080a192f74d2")
    rw = RandomWords()
    tab.set.window.max()
    chrome.wait(2, 3)

    if tab.ele('t:button@text():Connect Wallet'):
        logger.info(f"{env.name}   链接钱包选择OKX")
        tab.ele('t:button@text():Connect Wallet').click()
        chrome.wait(3, 6)
        tab.ele('OKX').click()
        chrome.wait(10, 15)

        if chrome.get_tab(title="OKX Wallet"):
            logger.info(f"{env.name}   OKX钱包授权")
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
            chrome.wait(10, 15)

    for _ in range(2):
        if tab.ele('t:button@text():Sign'):
            logger.info(f"{env.name}   登录")
            tab.ele('t:button@text():Sign').click()
            chrome.wait(5, 10)
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
            chrome.wait(3, 6)

    logger.info(f"{env.name}   进入任务页面")
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()

    tab.ele('t:span@text():START MISSIONS').click()
    chrome.wait(2, 3)

    try:

        if tab.ele('t:button@text():Authorize'):
            tab.ele('t:button@text():Authorize').click()
            chrome.wait(15, 25)
            chrome.get_tab(url='https://discord.com').ele("@type=button", index=2).click()
            chrome.wait(5, 10)
            logger.info(f"{env.name}    开始验证")


        count = 0
        while count < 2:
                element = tab.ele('t:button@text():Start')
                if not element:
                    break
                else:
                    element.click()
                count += 1

        try:
            tab.ele('t:button@text():Verify').click()
        except Exception as e:
            pass

        count = 0
        while count < 4:
                element = tab.ele('t:button@text():Start')
                if not element:
                    break
                else:
                    element.click()
                count += 1

    except Exception as e:
        pass

    tab = chrome.new_tab(url="https://dashboard.arch.network/missions")
    if tab.ele('t:button@text():Authorize') or tab.ele('t:button@text():Start') or tab.ele('t:button@text():Verify'):
        logger.info(f"{env.name}   有未完成任务！！！")

    return

def missions1(chrome, env):

    tab = chrome.new_tab(url="https://x.com/ArchNtwrk")
    rw = RandomWords()
    outlook: Account = Account.query.filter_by(id=env.outlook_id).first()
    random_word = rw.random_word()
    logger.info(f"{env.name}   开始做入职任务")
    chrome.wait(15, 25)
    tw_tab = chrome.get_tab(url="x.com")
    try:
        if tw_tab:
                if "login" in tw_tab.url:
                        logger.info(f"{env.name}: 推特未登录,尝试重新登录")
                        with app.app_context():
                            tw: Account = Account.query.filter_by(id=env.tw_id).first()
                        if tw:
                                tw_tab.ele("@autocomplete=username").input(tw.name, clear=True)
                                tw_tab.ele("@@type=button@@text()=Next").click()
                                tab.ele("t:span@text():Password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd), clear=True)
                                tw_tab.ele("@@type=button@@text()=Log in").click()
                                fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                                if "login" in tab.url and len(fa2) > 10:
                                    tw2faV(tab, fa2)
                                chrome.wait(25, 30)
                        else:
                                raise Exception(f"{env.name}: 没有导入TW的账号信息")

        if chrome.get_tab(url='https://x.com/'):
                        chrome.get_tab(url='https://x.com/').ele("t:span@text():Log in").click()
                        logger.info(f"{env.name}: 推特未登录,尝试重新登录")
                        with app.app_context():
                            tw: Account = Account.query.filter_by(id=env.tw_id).first()
                            if tw:
                                tw_tab.ele("@autocomplete=username").input(tw.name, clear=True)
                                tw_tab.ele("@@type=button@@text()=Next").click()
                                tab.ele("t:span@text():Password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd), clear=True)
                                tw_tab.ele("@@type=button@@text()=Log in").click()
                                fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                                if "login" in tab.url and len(fa2) > 10:
                                    tw2faV(tab, fa2)
                                chrome.wait(25, 30)
                            else:
                                raise Exception(f"{env.name}: 没有导入TW的账号信息")
    except Exception as e:
            pass

    try:
        if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Start"):
            chrome.get_tab(url='https://x.com/').ele("@@type=submit@@value=Start").click()
            chrome.wait(10, 15)
            if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Send email"):
                logger.info(f"{env.name}   该环境推特需要邮箱验证，请前往验证")
                chrome.wait(30, 40)
    except Exception as e:
        pass

    try:
        if chrome.get_tab(url='https://x.com/').ele("@@type=submit@@value=Continue to X"):
            chrome.get_tab(url='https://x.com/').ele("@@type=submit@@value=Continue to X").click()
            chrome.wait(20, 25)
        if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Start"):
            chrome.get_tab(url='https://twitter.com/').ele("@@type=submit@@value=Start").click()
            chrome.wait(10, 15)
        if chrome.get_tab(url='https://x.com/').s_ele("@@type=submit@@value=Send email"):
            logger.info(f"{env.name}   该环境推特需要邮箱验证，请前往验证")
            quitChrome(env, chrome)
    except Exception as e:
        pass

    try:
        chrome.get_tab(url='https://x.com/').ele("t:span@text():Follow").click()
        chrome.wait(2, 3)
    except Exception as e:
        pass

    if tab.ele('t:span@text():Refresh'):
        tab.ele('t:span@text():Refresh').click()
        chrome.wait(25, 30)
        tw_tab = chrome.get_tab(url="x.com")
        if tw_tab:
            if "login" in tw_tab.url:
                    logger.info(f"{env.name}: 推特未登录,尝试重新登录")
                    with app.app_context():
                        tw: Account = Account.query.filter_by(id=env.tw_id).first()

                    if tw:
                        tw_tab.ele("@autocomplete=username").input(tw.name, clear=True)
                        tw_tab.ele("@@type=button@@text()=Next").click()
                        tab.ele("t:span@text():Password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd), clear=True)
                        tw_tab.ele("@@type=button@@text()=Log in").click()
                        fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                        if "login" in tab.url and len(fa2) > 10:
                            tw2faV(tab, fa2)
                        chrome.wait(25, 30)
                    else:
                        raise Exception(f"{env.name}: 没有导入TW的账号信息")

    tab.close()
    tab = chrome.new_tab(url="https://dashboard.arch.network/missions")
    chrome.wait(2, 3)
    tab.ele('t:div@text():ONBOARDING MISSIONS').click()
    chrome.wait(2, 3)
    tab.ele('t:button@text():Start').click()
    chrome.wait(3, 6)

    tw_tab = chrome.get_tab(url="x.com")
    try:
        if tw_tab:
            if "login" in tw_tab.url:
                    logger.info(f"{env.name}: 推特未登录,尝试重新登录")
                    with app.app_context():
                        tw: Account = Account.query.filter_by(id=env.tw_id).first()
                    if tw:
                            tw_tab.ele("@autocomplete=username").input(tw.name, clear=True)
                            tw_tab.ele("@@type=button@@text()=Next").click()
                            tab.ele("t:span@text():Password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd), clear=True)
                            tw_tab.ele("@@type=button@@text()=Log in").click()
                            fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                            if "login" in tab.url and len(fa2) > 10:
                                tw2faV(tab, fa2)
                            chrome.wait(25, 30)
                    else:
                            raise Exception(f"{env.name}: 没有导入TW的账号信息")

        if chrome.get_tab(url='https://x.com/'):
                    chrome.get_tab(url='https://x.com/').ele("t:span@text():Log in").click()
                    logger.info(f"{env.name}: 推特未登录,尝试重新登录")
                    with app.app_context():
                        tw: Account = Account.query.filter_by(id=env.tw_id).first()
                        if tw:
                            tw_tab.ele("@autocomplete=username").input(tw.name, clear=True)
                            tw_tab.ele("@@type=button@@text()=Next").click()
                            tab.ele("t:span@text():Password").input(aesCbcPbkdf2DecryptFromBase64(tw.pwd), clear=True)
                            tw_tab.ele("@@type=button@@text()=Log in").click()
                            fa2 = aesCbcPbkdf2DecryptFromBase64(tw.fa2)
                            if "login" in tab.url and len(fa2) > 10:
                                tw2faV(tab, fa2)
                            chrome.wait(25, 30)
                        else:
                            raise Exception(f"{env.name}: 没有导入TW的账号信息")
    except Exception as e:
        pass
    tab.close()

    try:
        tab = chrome.new_tab(url="https://x.com/compose/post")
        chrome.wait(5, 10)
        if tab.ele('t:span@text():Log in'):
            logger.info(f"{env.name}    推特登录失败")
            quitChrome(env, chrome)

        tab.ele("@class=css-175oi2r r-1iusvr4 r-16y2uox r-1777fci r-1h8ys4a r-1bylmt5 r-13tjlyg r-7qyjyx r-1ftll1t").input('Want to earn points and help build bridgeless Bitcoin DeFi?\n\nJoin me: https://dashboard.arch.network?referralCode=254c68f9-1b28-4e66-bdf1-4e42c48085f9 #JoinArch ')
        chrome.wait(3, 6)
        tab.ele("t:span@text():Post").click(by_js=True)
        chrome.wait(3, 6)
        tab.close()

        tab = chrome.new_tab(url="https://x.com/ArchNtwrk")
        try:
            chrome.get_tab(url='https://x.com/').ele("t:span@text():Follow").click()
            chrome.wait(3, 6)
            chrome.get_tab(url='https://x.com/').close()
        except Exception as e:
            pass
    except Exception as e:
        pass

    tab = chrome.new_tab(url="https://www.youtube.com/channel/UCSsjwRKAUnCb6sj38wk0YvQ")
    chrome.wait(3, 6)
    tab.close()
    tab = chrome.new_tab(url="https://dashboard.arch.network/missions")

    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    try:
        tab.ele('t:div@text():NOTIFICATIONS').click()
        chrome.wait(2, 3)
        tab.ele('t:div@text():Next').click()
        chrome.wait(2, 3)
    except Exception as e:
        pass
    if chrome.get_tab(title="OKX Wallet"):
        logger.info(f"{env.name}   OKX钱包授权")
        chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
        chrome.wait(10, 15)
    if tab.ele('@placeholder=Email address'):
        tab.ele('@placeholder=Email address').input(outlook.name, clear=True)
        chrome.wait(2, 3)
    if tab.ele('@class=notifi-ftu-target-edit-button-text'):
        tab.ele('@class=notifi-ftu-target-edit-button-text').click()
        chrome.wait(3, 6)
    if tab.ele('@class=btn notifi-ftu-target-list-button'):
        tab.ele('@class=btn notifi-ftu-target-list-button').click()
        chrome.wait(3, 6)
    if tab.ele('@class=btn notifi-ftu-alert-edit-button'):
        tab.ele('@class=btn notifi-ftu-alert-edit-button').click()

    tab.close()
    tab = chrome.new_tab(url="https://dashboard.arch.network/missions")
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    tab.ele('t:div@text():ONBOARDING MISSIONS').click()
    chrome.wait(2, 3)


    logger.info(f"{env.name}    开始验证")
    try:
        count = 0
        while count < 6:
            element = tab.ele('t:button@text():Start')
            if not element:
                break
            else:
                element.click()
            count += 1
        chrome.wait(5, 10)

        if tab.ele('t:span@text():CONTINUE'):
            tab.ele('t:span@text():CONTINUE').click()
        if tab.ele('t:span@text():START MISSIONS'):
            tab.ele('t:span@text():START MISSIONS').click()
            chrome.wait(2, 3)
        if tab.ele('t:span@text():CONTINUE'):
            tab.ele('t:span@text():CONTINUE').click()
        chrome.wait(2, 4)

        chrome.get_tab(url='https://x.com/').ele("t:span@text():Post").click(by_js=True)
        chrome.wait(2, 4)
        chrome.get_tab(url='https://x.com/').close()
        tab.close()
        tab = chrome.new_tab(url="https://dashboard.arch.network/missions")
    except Exception as e:
        pass


    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
        chrome.wait(2, 3)
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    tab.ele('t:div@text():ONBOARDING MISSIONS').click()

    chrome.wait(2, 3)
    count = 0
    while count < 6:
        element = tab.ele('t:button@text():Start')
        if not element:
            break
        else:
            element.click()
        count += 1

    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
        chrome.wait(2, 3)
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()

    tab.close()
    return

def weekly(chrome, env):
    logger.info(f"{env.name}    开始做每周任务")
    tab = chrome.new_tab(url="https://dashboard.arch.network?referralCode=f9c6ab90-03a4-4724-9cbd-080a192f74d2")
    rw = RandomWords()
    tab.set.window.max()

    if tab.ele('t:button@text():Connect Wallet'):
        logger.info(f"{env.name}   链接钱包选择OKX")
        tab.ele('t:button@text():Connect Wallet').click()
        chrome.wait(3, 6)
        tab.ele('OKX').click()
        chrome.wait(10, 15)

        if chrome.get_tab(title="OKX Wallet"):
            logger.info(f"{env.name}   OKX钱包授权")
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
            chrome.wait(10, 15)

    for _ in range(2):
        if tab.ele('t:button@text():Sign'):
            logger.info(f"{env.name}   登录")
            tab.ele('t:button@text():Sign').click()
            chrome.wait(3, 6)
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
            chrome.wait(3, 6)

    logger.info(f"{env.name}   进入任务页面")
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()

    tab.ele('t:span@text():START MISSIONS').click()
    chrome.wait(2, 3)
    tab.ele('t:div@text():WEEKLY MISSIONS').click()

    x = chrome.new_tab(url="https://x.com/predictr_market")
    x.ele('t:span@text():Follow').click()
    chrome.wait(1, 2)

    # tab.ele('t:button@text():Start').click()
    # chrome.wait(1)
    # if tab.ele('t:div@text():Completed'):
    #     pass
    # else:
    #     logger.info(f"{env.name}   推特关注失败")

    count = 0
    while count < 2:
        element = tab.ele('t:button@text():Start')
        if not element:
            break
        else:
            element.click()
        count += 1
    chrome.wait(3, 6)

    if tab.ele('t:div@text():Completed'):
        pass
    else:
        logger.info(f"{env.name}   推特关注失败")

    return


def daily(chrome, env):
    logger.info(f"{env.name}   开始做每日任务")
    tab = chrome.new_tab(url="https://dashboard.arch.network/missions")
    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
        chrome.wait(2, 3)
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
        chrome.wait(2, 3)
    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
    chrome.wait(2, 3)
    tab.ele('t:div@text():DAILY MISSIONS').click()
    chrome.wait(2, 3)
    count = 0
    while count < 14:
        element = tab.ele('t:button@text():Start')
        if not element:
            break
        else:
            element.click()
        count += 1
    chrome.wait(5, 10)
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
        chrome.wait(2, 3)
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()

    tab.close()
    return


def community(chrome, env):
    logger.info(f"{env.name}   开始做社区任务")
    tab = chrome.new_tab(url="https://x.com/Saturn_btc")
    try:
        chrome.get_tab(url='https://x.com/').ele("t:span@text():Follow").click()
        chrome.wait(3, 6)
        chrome.get_tab(url='https://x.com/').close()
    except Exception as e:
        pass

    tab = chrome.new_tab(url="https://x.com/bimabtc")
    try:
        chrome.get_tab(url='https://x.com/').ele("t:span@text():Follow").click()
        chrome.wait(3, 6)
        chrome.get_tab(url='https://x.com/').close()
    except Exception as e:
        pass

    tab = chrome.new_tab(url="https://x.com/BoundUSD")
    try:
        chrome.get_tab(url='https://x.com/').ele("t:span@text():Follow").click()
        chrome.wait(3, 6)
        chrome.get_tab(url='https://x.com/').close()
    except Exception as e:
        pass

    tab = chrome.new_tab(url="https://x.com/funkybit_fun")
    try:
        chrome.get_tab(url='https://x.com/').ele("t:span@text():Follow").click()
        chrome.wait(3, 6)
        chrome.get_tab(url='https://x.com/').close()
    except Exception as e:
        pass

    tab = chrome.new_tab(url="https://dashboard.arch.network/missions")
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()

    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
        chrome.wait(2, 3)
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    chrome.wait(2, 3)

    tab.ele('t:div@text():COMMUNITY MISSIONS').click()

    chrome.wait(2, 3)
    count = 0
    while count < 8:
        element = tab.ele('t:button@text():Start')
        if not element:
            break
        else:
            element.click()
        count += 1
    chrome.wait(5, 10)

    tab.close()
    return

def faucet(chrome, env):
    tab = chrome.new_tab(url="https://faucet.bound.money/")

    # 打开文件进行读取和写入
    while True:

        # 以读取模式打开文件并读取一行
        with open(r'C:\Users\Toka\Desktop\shell\wallet.txt', 'r', encoding='utf-8') as file:
            lines = file.readlines()

        # 如果文件为空，跳出循环
        if not lines:
            break

        # 获取文件的第一行
        line = lines[0].strip()  # 读取并去掉行尾的换行符

        if wait_captcha_page(tab, env):
            # 执行操作
            tab.ele('@type=text').input(line, clear=True)
            chrome.wait(5)
            tab.ele('@type=submit').click()
            chrome.wait(1, 2)
            if tab.ele('t:div@text=Limit exceeded') or tab.ele('t:p@text():Server is overload') or tab.ele('t:p@text()=Invalid capcha') or tab.ele('t:div@text():Try again later') or tab.ele('Try again later'):
                with open(r'C:\Users\Toka\Desktop\shell\wallet.txt', 'w', encoding='utf-8') as file:
                    file.writelines(lines)
                logger.info(f"{env.name}   该环境已达到限制，暂时关闭")
                quitChrome(env, chrome)

            if tab.ele('t:div@text=Limit exceeded') or tab.ele('t:p@text():Server is overload') or tab.ele('t:p@text()=Invalid capcha') or tab.ele('t:div@text():Try again later') or tab.ele('Try again later'):
                with open(r'C:\Users\Toka\Desktop\shell\wallet.txt', 'w', encoding='utf-8') as file:
                    file.writelines(lines)
                logger.info(f"{env.name}   该环境已达到限制，暂时关闭")
                quitChrome(env, chrome)

            # 每处理完一行后，删除这一行
            lines.pop(0)

            # 将剩余内容重新写回文件
            with open(r'C:\Users\Toka\Desktop\shell\wallet.txt', 'w', encoding='utf-8') as file:
                file.writelines(lines)
        else:
            logger.info(f"{env.name}   该环境已达到限制，暂时关闭")
            quitChrome(env, chrome)

    return


def faucet2(chrome, env):
    tab = chrome.new_tab(url="https://faucet.testnet4.dev/")

        # 以读取模式打开文件并读取一行
    with open(r'C:\Users\Toka\Desktop\shell\wallet.txt', 'r', encoding='utf-8') as file:
        lines = file.readlines()

        # 如果文件为空，跳出循环
        if not lines:
            quitChrome(env, chrome)

        # 获取文件的第一行
    line = lines[0].strip()  # 读取并去掉行尾的换行符

    if wait_captcha_page(tab, env):
            # 执行操作
            tab.ele('@type=text').input(line, clear=True)
            chrome.wait(5)
            tab.ele('@type=submit').click()
            chrome.wait(1, 2)
            # 每处理完一行后，删除这一行
            lines.pop(0)

            # 将剩余内容重新写回文件
            with open(r'C:\Users\Toka\Desktop\shell\wallet.txt', 'w', encoding='utf-8') as file:
                file.writelines(lines)

    else:
            logger.info(f"{env.name}   该环境已达到限制，暂时关闭")
            quitChrome(env, chrome)

    return

import openpyxl
import time
def count(chrome, env):
    tab = chrome.new_tab(url="https://dashboard.arch.network/missions")
    taskData = getTaskObject(env, name)
    env_name = env.name
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    if tab.ele('t:span@text():START MISSIONS'):
        tab.ele('t:span@text():START MISSIONS').click()
    if tab.ele('t:span@text():CONTINUE'):
        tab.ele('t:span@text():CONTINUE').click()
    chrome.wait(2, 3)

    logger.info(f"{env.name}   开始数据统计")

    xp = tab.ele('@class=absolute inset-0 flex items-center justify-center text-[18px] font-bold leading-normal text-lighter-yellow', index=1).text
    chrome.wait(2, 4)
    level = tab.ele('@class=text-lightest-yellow text-[15px] leading-6 uppercase').text

        # 使用原始字符串方式指定文件路径
    current_time = time.strftime("%m-%d")
    file_path = r'C:\Users\Public\Documents\arch_{}.xlsx'.format(current_time)

        # 打开已存在的 Excel 文件（arch.xlsx）
    try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = 'XP'
            ws['C1'] = 'Level'
    except FileNotFoundError:
            # 如果文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = 'XP'
            ws['C1'] = 'Level'
            wb.save(file_path)
        # 找到下一行位置（避免覆盖）
    next_row = ws.max_row + 1

    env_name_exists = False
    for row in range(2, ws.max_row + 1):  # 从第二行开始遍历（跳过表头）
        if ws[f'A{row}'].value == env_name:
                # 如果找到相同的 env_name，更新该行的 xp 和 level
                ws[f'B{row}'] = xp
                ws[f'C{row}'] = level
                env_name_exists = True
                break
    if not env_name_exists:
            # 如果没有找到相同的 env_name，追加新行
            next_row = ws.max_row + 1
            ws[f'A{next_row}'] = env_name
            ws[f'B{next_row}'] = xp
            ws[f'C{next_row}'] = level

        # 保存文件（不会覆盖，直接追加）
    wb.save(file_path)

        # 更新任务记录
    taskData.Xp = xp
    taskData.Level = level
    updateTaskRecord(env.name, name, taskData, 1)

    return


def arch(env):
    with app.app_context():
        try:
            chrome: ChromiumPage = OKXChrome(env)
            # getTab(chrome, env)
            # missions(chrome, env)
            weekly(chrome, env)
            # daily(chrome, env)
            # community(chrome, env)
            # count(chrome, env)
            # faucet(chrome, env)
            # faucet2(chrome, env)
            logger.info(f"{env.name}环境：任务执行完毕，关闭环境")
        except Exception as e:
            logger.error(f"{env.name} 执行：{e}")
            return ("失败", e)
        finally:
            quitChrome(env, chrome)






