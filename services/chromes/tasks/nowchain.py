import openpyxl
from DrissionPage import ChromiumPage,ChromiumOptions
from loguru import logger
# 连接数据库
from flaskServer.config.connect import app
#登录环境账号
from flaskServer.services.chromes.login import OKXChrome
from flaskServer.services.dto.task_record import updateTaskRecord, getTaskObject
import time
import random
from flaskServer.utils.chrome import quitChrome

#项目名称
name = 'NowChain'
#项目邀请链接
now_chain_url = 'https://testnet.nowchain.co/testnet/point-system?referral=0xECB41b49D74D7d13bB51f9603Fd2360557647504/'
switch_Network = '''let button  =
document.querySelector("body > w3m-modal").shadowRoot.querySelector("wui-flex > wui-card > w3m-router").shadowRoot.querySelector("div > w3m-unsupported-chain-view").shadowRoot.querySelector("wui-flex > wui-flex:nth-child(2) > wui-list-network:nth-child(2)");
button.click();
'''
okx_url = '''let button  =
document.querySelector("body > w3m-modal").shadowRoot.querySelector("wui-flex > wui-card > w3m-router").shadowRoot.querySelector("div > w3m-connect-view").shadowRoot.querySelector("wui-flex > w3m-wallet-login-list").shadowRoot.querySelector("wui-flex > w3m-connect-announced-widget").shadowRoot.querySelector("wui-flex > wui-list-wallet:nth-child(1)");
button.click();'''

def exe_okx(chrome,env):
    try:
        if chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2):
            chrome.wait(3, 4)
            chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2).click()
            chrome.wait(3, 4)
            try:
                chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2).click()
            except Exception as e:
                pass
    except Exception as e:
        pass

#主页面登录
def getTab(chrome,env):
    tab = chrome.new_tab(now_chain_url)
    # 设置全屏
    tab.set.window.max()
    tab.wait.load_start(timeout=6)
    try:
        tab.run_js(switch_Network)
        print('选择测试网完成')
        tab.wait.load_start(timeout=3)
        exe_okx(chrome, env)
    except Exception as e:
        pass
    try:
        # print('开始选择钱包并且授权')
        tab.wait(5).run_async_js(okx_url)
        tab.wait.load_start(timeout=3)
        exe_okx(chrome, env)
    except Exception as e:
        pass
    #登录钱包
    if tab.wait.ele_displayed('Connect Wallet', timeout=5, raise_err=False):
        print('开始链接钱包')
        tab.wait.load_start(timeout=6)
        tab.ele('Connect Wallet').click()
    else:
        print('不需要Connect Wallet按钮')
    try:
        print('开始选择钱包并且授权')
        tab.wait(5).run_async_js(okx_url)
        tab.wait.load_start(timeout=3)
        exe_okx(chrome, env)
    except Exception as e:
        pass
    try:
        tab.run_js(switch_Network)
        print('选择测试网完成')
        tab.wait.load_start(timeout=3)
        exe_okx(chrome, env)
    except Exception as e:
        pass
#领水
def getFaucet(chrome, env):
    getTab(chrome, env)
    time.sleep(5)
    print('开始领水')
    tab = chrome.new_tab(url='https://testnet.nowchain.co/testnet/faucet/')
    try:
        if tab.wait.eles_loaded('Time remaining: ', timeout=10, raise_err=False):
            print('领水时间还没到：', tab.ele('Time remaining: ').text)
            return False
        tab.wait.load_start(timeout=6)
        num = 0
        while num < 5:
            if tab.wait.eles_loaded('Request Assets', timeout=10, raise_err=False):
                print('此时为Request Assets，开始点击领水')
                tab.ele('t:button@tx():Request Assets').click()
                print('点击完成')
                num += 1
                tab.wait.load_start(timeout=6)
            else:
                print('领水完成')
                return True
    except Exception as e:
        logger.error(e)
        return False

#check——in
def getChck_in(chrome,env):
    # getTab(chrome, env)
    tab = chrome.new_tab(url=now_chain_url)
    time.sleep(5)
    chrome.refresh(ignore_cache= True)
    time.sleep(5)
    #登录钱包
    try:
        if tab.s_ele('t:button@tx():Checked'):
            logger.info('已经签到完成,或者还没有到签到时间')
            check_day = tab.ele('@class=px-4 text-sm sm:text-base py-1.5 rounded-lg text-primary-10 bg-BG-5 font-semibold').text
            return check_day
        elif tab.s_ele('t:button@tx():Check-in'):
            print('开始签到')
            tab.wait.ele_displayed('t:button@tx():Check-in', timeout=50)
            tab.ele('t:button@tx():Check-in').click()
            time.sleep(5)
            exe_okx(chrome,env)
            time.sleep(20)
            chrome.refresh(ignore_cache=True)
            check = 0
            while check < 3:
                if tab.s_ele('t:button@tx():Checked'):
                    logger.info('已经签到完成,或者还没有到签到时间')
                    check_day = tab.ele('@class=px-4 text-sm sm:text-base py-1.5 rounded-lg text-primary-10 bg-BG-5 font-semibold').text
                    return check_day
                elif tab.s_ele('t:button@tx():Check-in'):
                    tab.wait.ele_displayed('t:button@tx():Check-in', timeout=50)
                    tab.ele('t:button@tx():Check-in').click()
                    time.sleep(5)
                    exe_okx(chrome,env)
                    time.sleep(20)
                    check += 1
                    chrome.refresh(ignore_cache=True)
                    if check == 3:
                        check_day = tab.ele('@class=px-4 text-sm sm:text-base py-1.5 rounded-lg text-primary-10 bg-BG-5 font-semibold').text
                        return check_day
    except Exception as e:
        logger.error(e)
        check_day = tab.ele('@class=px-4 text-sm sm:text-base py-1.5 rounded-lg text-primary-10 bg-BG-5 font-semibold').text
        return check_day

#swap
def getSwap(chrome,env):
    tab = chrome.new_tab(url='https://testnet.nowchain.co/testnet/swap/')
    time.sleep(5)
    try:
        # 在0.001~0.05随机一个数，保留小数点后四位的浮点数
        random_number = round(random.uniform(0.0001, 0.01), 4)
        print('随机数是：',random_number)
        tab.wait.ele_displayed('@class=MuiInputBase-input css-mnn31', timeout=20)
        tab.ele('@class=MuiInputBase-input css-mnn31').input(random_number)
        if tab.s_ele('Swap Now'):
            print('开始输入随机数')
            tab.wait.ele_displayed('Swap Now', timeout=20)
            print('开始确认swap')
            tab.ele('Swap Now').click()
            time.sleep(5)
            exe_okx(chrome,env)
            time.sleep(5)
            return True
        elif tab.s_ele('Insufficient balance'):
            print('余额低于预期，只能再降')
            tab.ele('@class=MuiInputBase-input css-mnn31').clear()
            tab.ele('@class=MuiInputBase-input css-mnn31').input(0.000001)
            tab.wait.ele_displayed('Swap Now', timeout=20)
            tab.ele('Swap Now').click()
            time.sleep(5)
            exe_okx(chrome,env)
            time.sleep(5)
            return True
    except Exception as e:
        logger.error(e)
        return False

#Liquidity
def getLiquidity(chrome,env):
    tab = chrome.new_tab(url='https://testnet.nowchain.co/testnet/liquidity/')
    time.sleep(5)
    try:
        # 在0.001~0.05随机一个数，保留小数点后两位的浮点数
        random_number = round(random.uniform(0.0001, 0.01), 4)
        print('随机数是：', random_number)
        tab.wait.ele_displayed('@class=MuiInputBase-input css-mnn31', timeout=20)
        print('开始输入随机数')
        tab.ele('@class=MuiInputBase-input css-mnn31').input(random_number)
        #开始选择兑换对象
        tab.wait.ele_displayed('Select Coin', timeout=10)
        tab.ele('Select Coin').click()
        time.sleep(5)
        #这里随机选择数字就可以兑换对应的测试币，注：2：BNB  3:USDT  4:ETH
        items= [
            '2',
            '3',
            '4'
        ]
        random_item = int(random.choice(items))
        print('随机的数是：', random_item)
        tab.ele('@class=flex items-center justify-start gap-1.5 sm:gap-2 w-full', index=random_item).click()
        time.sleep(5)
        if tab.s_ele('Add Pair'):
            tab.wait.ele_displayed('Add Pair', timeout=10)
            tab.ele('Add Pair').click()
            time.sleep(5)
            exe_okx(chrome,env)
            time.sleep(5)
            if tab.s_ele('Enter amount'):
                return True
        elif tab.s_ele('Insufficient balance'):
            print('余额低于预期，只能再降')
            tab.ele('@class=MuiInputBase-input css-mnn31').clear()
            tab.ele('@class=MuiInputBase-input css-mnn31').input(0.000001)
            tab.wait.ele_displayed('Add Pair', timeout=10)
            tab.ele('Add Pair').click()
            time.sleep(5)
            exe_okx(chrome,env)
            time.sleep(5)
            if tab.s_ele('Enter amount'):
                return True
    except Exception as e:
        logger.error(e)
        return False

#Bridge
def getBridge(chrome,env):
    tab = chrome.new_tab(url='https://testnet.nowchain.co/testnet/bridge/')
    time.sleep(5)
    try:
        random_number = round(random.uniform(0.0001, 0.01), 4)
        print('随机数是：',random_number)
        tab.wait.ele_displayed('@class=MuiInputBase-input css-mnn31', timeout=20)
        tab.ele('@class=MuiInputBase-input css-mnn31').input(random_number)
        if tab.s_ele('Time remaining: '):
            print('已经质押或者还没到时间：',tab.ele('Time remaining: ').text)
            return False
        elif tab.s_ele('Bridge Now'):
            tab.wait.ele_displayed('Bridge Now', timeout=20)
            tab.ele('Bridge Now').click()
            time.sleep(5)
            exe_okx(chrome,env)
            time.sleep(5)
            if tab.s_ele('Enter amount'):
                return True
        elif tab.s_ele('Insufficient balance'):
            print('余额低于预期，只能再降')
            tab.ele('@class=MuiInputBase-input css-mnn31').clear()
            tab.ele('@class=MuiInputBase-input css-mnn31').input(0.000001)
            tab.wait.ele_displayed('Bridge Now', timeout=20)
            tab.ele('Bridge Now').click()
            time.sleep(5)
            exe_okx(chrome,env)
            time.sleep(5)
            if tab.s_ele('Enter amount'):
                return True
    except Exception as e:
        logger.error(e)
        return False

def getYesCaptchaassistant(chrome,env):
    tab = chrome.new_tab(url='chrome-extension://phnemkgfgnkkpagdlpccniemhdmogbah/popup/index.html')
    time.sleep(5)
    try:
        if tab.s_ele('@class=MuiButtonBase-root MuiSwitch-switchBase MuiSwitch-colorDefault PrivateSwitchBase-root MuiSwitch-switchBase MuiSwitch-colorDefault css-1g3pnk5'):
            tab.ele('@class=MuiButtonBase-root MuiSwitch-switchBase MuiSwitch-colorDefault PrivateSwitchBase-root MuiSwitch-switchBase MuiSwitch-colorDefault css-1g3pnk5').click()
            print('点击完成')
            time.sleep(5)
        elif tab.s_ele('@class=MuiButtonBase-root MuiSwitch-switchBase MuiSwitch-colorDefault Mui-checked PrivateSwitchBase-root MuiSwitch-switchBase MuiSwitch-colorDefault Mui-checked Mui-checked css-1g3pnk5'):
            print('开关已经打开')
            return
    except Exception as e:
        logger.error(e)
#统计数量
def getCount(chrome, env):
    try:
        taskData = getTaskObject(env, name)
        tab = chrome.new_tab(now_chain_url)
        time.sleep(10)

        # 统计总数
        try:
            print('统计总数')
            PointsCount = tab.ele('@class=flex items-center gap-3 text-base sm:text-lg font-semibold').text
            PointsCount_num = PointsCount.split('Points')[0]
            print(f'{env.name}的PointsCount_num:', PointsCount_num)
            print('开始上传总数')
            taskData.PointsCount = PointsCount_num
            print('上传总数完成')
        except Exception as e:
            logger.error(e)
            PointsCount = tab.ele('@class=p-4 sm:p-5 rounded-lg sm:rounded-xl shadow flex justify-between gap-5 bg-BG-4').text
            PointsCount_num = PointsCount.split('Points')[1].strip()
            print(f'{env.name}的p2s:', PointsCount_num)
            print('开始上传总数')
            taskData.PointsCount = PointsCount_num
            print('上传总数完成')

        time.sleep(5)
        Faucet = getFaucet(chrome, env)
        print(f'{env.name}的Faucet:', Faucet)
        time.sleep(5)

        getChck_in(chrome, env)
        chrome.close_tabs()
        # # 统计签到天数
        check_in = tab.ele('@class=px-4 text-sm sm:text-base py-1.5 rounded-lg text-primary-10 bg-BG-5 font-semibold').text
        check_in_num = check_in.split('days')[0]
        print(f'{env.name}的check_in_num:', check_in_num)

        Swap = getSwap(chrome, env)
        print(f'{env.name}的Swap:', Swap)
        time.sleep(5)
        Bridge = getBridge(chrome, env)
        print(f'{env.name}的Bridge:', Bridge)
        time.sleep(5)
        Liquidity = getLiquidity(chrome, env)
        print(f'{env.name}的Liquidity:', Liquidity)
        time.sleep(5)

        print('开始上传数据：')
        taskData.check_in = check_in_num
        print('check_in_num：',check_in_num)
        taskData.Faucet = Faucet
        print('Faucet:',Faucet)
        taskData.Swap = Swap
        print('Swap:',Swap)
        taskData.Bridge = Bridge
        print('Bridge:',Bridge)
        taskData.Liquidity = Liquidity
        print('Liquidity:',Liquidity)
        taskData.Leaderboard = 0
        updateTaskRecord(env.name, name, taskData, 1)
        time.sleep(10)

        #存储数据到桌面
        current_time = time.strftime("%m-%d")
        file_path = r'C:\Users\Public\Documents\nowchain_{}.xlsx'.format(current_time)

        # 打开已存在的 Excel 文件（arch.xlsx）
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = 'check_in'
            ws['C1'] = 'PointsCount'
            ws['D1'] = 'Liquidity'
            ws['E1'] = 'Faucet'
            ws['F1'] = 'Bridge'
        except FileNotFoundError:
            # 如果文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = 'check_in'
            ws['C1'] = 'PointsCount'
            ws['D1'] = 'Liquidity'
            ws['E1'] = 'Faucet'
            ws['F1'] = 'Bridge'
            wb.save(file_path)
        # 找到下一行位置（避免覆盖）
        next_row = ws.max_row + 1

        env_name_exists = False
        env_name = env.name
        for row in range(2, ws.max_row + 1):  # 从第二行开始遍历（跳过表头）
            if ws[f'A{row}'].value == env_name:
                # 如果找到相同的 env_name，更新该行的 xp 和 level
                ws[f'B{row}'] = check_in
                ws[f'C{row}'] = PointsCount
                ws[f'D{row}'] = Liquidity
                ws[f'E{row}'] = Faucet
                ws[f'F{row}'] = Bridge

                env_name_exists = True
                break
        if not env_name_exists:
            # 如果没有找到相同的 env_name，追加新行
            next_row = ws.max_row + 1
            ws[f'A{next_row}'] = env_name
            ws[f'B{next_row}'] = check_in
            ws[f'C{next_row}'] = PointsCount
            ws[f'D{next_row}'] = Liquidity
            ws[f'E{next_row}'] = Faucet
            ws[f'F{next_row}'] = Bridge

        # 保存文件（不会覆盖，直接追加）
        wb.save(file_path)
    except Exception as e:
        logger.error(e)


def NowChain(env):
    with app.app_context():
        try:
            chrome: ChromiumPage = OKXChrome(env)
            getTab(chrome,env)
            chrome.close_tabs()
            getCount(chrome, env)

            logger.info(f"{env.name}环境：任务执行完毕，关闭环境")
        except Exception as e:
            logger.error(f"{env.name} 执行：{e}")
            return ("失败", e)
        finally:
            quitChrome(env, chrome)