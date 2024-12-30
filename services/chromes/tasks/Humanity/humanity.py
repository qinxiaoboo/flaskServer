import string
import multiprocessing
import openpyxl
import requests
from DrissionPage import ChromiumPage,ChromiumOptions
from loguru import logger
# 连接数据库
from flaskServer.config.connect import app
from flaskServer.mode.account import Account
#登录环境账号
from flaskServer.services.chromes.login import OKXChrome
from flaskServer.services.dto.account import updateAccountStatus
from flaskServer.services.dto.task_record import updateTaskRecord, getTaskObject
import time
import random
from flaskServer.utils.chrome import quitChrome, get_Custome_Tab
from flaskServer.utils.crypt import aesCbcPbkdf2DecryptFromBase64
from flaskServer.utils.RedisHelp import RedisLock

#humanity protocol
name = 'humanity'
humanity_url = 'https://testnet.humanity.org/login?ref=269toka'

wallet_js = '''
let button  = 
document.querySelector("body > div:nth-child(31) > div > div > div._9pm4ki5.ju367va.ju367v15.ju367v8r > div > div > div > div > div > div.iekbcc0.ju367v6p._1vwt0cg2.ju367v7a.ju367v7v > div.iekbcc0.ju367va.ju367v15.ju367v1n > div:nth-child(1) > button > div > div");
button.click();
'''
dis_js = '''
let button  = 
document.querySelector("body > div.MuiBox-root.mui-186aokm > div > div > div > div.MuiBox-root.mui-1i3hx1p > div > div > div.MuiBox-root.mui-17age7i > a:nth-child(2)");
button.click();
'''

def exe_okx(chrome,env):
    try:
        if chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2):
            chrome.wait(3, 4)
            chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2).click()
            chrome.wait(3, 4)
            try:
                chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2).click()
            except Exception as e:
                print('不需要二次确认')
    except Exception as e:
        print(f'{env.name}取的ele不对或者不需要连接')

def generate_random_word(length=7):
    # 生成一个随机的字母单词
    letters = string.ascii_lowercase + string.digits  # 包含小写字母、大写字母和数字
    return ''.join(random.choice(letters) for _ in range(length))

def getDiscord(chrome,env):
    print('开始discord认证')

    try:
        if chrome.get_tab(url='https://discord.com').ele("ok") or chrome.get_tab(url='https://discord.com').ele("好的"):
            try:
                chrome.get_tab(url='https://discord.com').ele("ok").click()
            except Exception as e:
                pass
            try:
                chrome.get_tab(url='https://discord.com').ele("好的").click()
            except Exception as e:
                pass

    except Exception as e:
        pass

    try:
        chrome.get_tab(url='https://discord.com').ele("@type=button", index=2).click(by_js=None)
        chrome.wait(5, 10)
    except Exception as e:
        logger.info(f"{env.name}   登录Discord失败！需要人工登录")
        quitChrome(env, chrome)

def gethumanity(chrome,env):
    tab = chrome.new_tab(url=humanity_url)
    tab.set.window.max()
    if tab.s_ele('Almost there!'):
        print('进入 Almost there! 等待状态')
        num = 0
        while num < 3:  # 最多尝试3次
            try:
                # 刷新页面
                chrome.refresh()
                # 等待页面加载
                if tab.wait.load_start(timeout=10, raise_err=False):
                    print("页面加载完成")

                if tab.s_ele('t:span@text():Get a new place in line'):
                    tab.ele('t:span@text():Get a new place in line').click()
                # 等待 'Get Started' 元素出现
                if tab.s_ele('Get Started', timeout=10):
                    print("'Get Started' 元素已加载，跳出循环")
                    break  # 如果 'Get Started' 元素出现，跳出循环
            except Exception as e:
                print(f"发生错误: {e}")
            num += 1  # 尝试次数加1
            if num < 3:
                # print(f"尝试次数 {num}, 重新刷新页面...")
                logger.info(f"尝试次数 {num}, 重新刷新页面...")

            else:
                logger.info("尝试3次刷新后仍未找到 'Get Started' 元素。")
                return
    else:
        print('不需要等待')
    # tab.wait.load_start(timeout=6)
    if tab.s_ele('t:p@text():Loading your profile...'):
        chrome.wait(15, 30)
        if tab.s_ele('t:p@text():Loading your profile...'):
            tab.refresh()
    else:
        tab.refresh()

    chrome.wait(15, 20)
    if tab.wait.ele_displayed('@class=skip', timeout=15, raise_err=False):
        print('点击skip弹幕')
        tab.ele('@class=skip').click()
    else:
        print('没有出现skip')
    #
    if tab.wait.ele_displayed('@class=skip', timeout=5, raise_err=False):
        print('点击skip弹幕')
        tab.ele('@class=skip').click()

    if tab.wait.ele_displayed('skip', timeout=10, raise_err=False):
        print('点击skip弹幕')
        tab.ele('@class=skip').click()

    if tab.wait.ele_displayed('@class=bottom', timeout=5, raise_err=False):
        print('点击签到')
        with RedisLock(f"{env.name}-okx", 200, 200):
            tab.ele('@class=bottom').click(by_js=None)
            chrome.wait(5, 8)
            try:
                if chrome.get_tab(title="OKX Wallet"):
                    chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
                    chrome.wait(15, 20)
            except Exception as e:
                pass


    elif tab.wait.ele_displayed('@class=bottom disable', timeout=5, raise_err=False):
        print('已经签到过不需要签到了')


    else:
        # tab.refresh()
        if tab.wait.eles_loaded('Get Started', timeout=5, raise_err=False):
            try:
                tab.ele('@class=MuiButtonBase-root MuiIconButton-root MuiIconButton-colorPrimary MuiIconButton-sizeMd mui-1o8rzlg', index=2).click()
                chrome.wait(15, 20)
            except Exception as e:
                pass

            try:
                tab.ele('Connect Wallet').click()
                chrome.wait(5, 10)
                if tab.s_ele('@data-testid=rk-auth-message-button'):
                    tab.ele('@data-testid=rk-auth-message-button').click()
                    chrome.wait(5, 10)
                try:
                    tab.ele('t:div@text():MetaMask').click()
                    chrome.wait(3, 6)
                except Exception as e:
                    pass
                if tab.s_ele('@data-testid=rk-auth-message-button'):
                    tab.ele('@data-testid=rk-auth-message-button').click()
                    chrome.wait(5, 10)
                chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
                chrome.wait(15, 20)
            except Exception as e:
                pass

        try:
            if chrome.get_tab(url='https://discord.com').ele('t:div@text():Log in') or chrome.get_tab(url='https://discord.com').ele('t:div@text():登录'):
                print('重新登录discord')
                chrome.wait(2, 3)
                updateAccountStatus(env.discord_id, 0, "重置了Discord登录状态")
                tab = chrome.get_tab(url="https://discord.com/")
                if tab.s_ele("Please log in again") or tab.s_ele("请再次登录"):
                    tab.ele(
                        "@class=button_dd4f85 lookFilled_dd4f85 colorPrimary_dd4f85 sizeMedium_dd4f85 grow_dd4f85").click()

                if "login" or "登录" in tab.url:
                    logger.info(f"{env.name} 开始登录 Discord 账号")
                    with app.app_context():
                        discord: Account = Account.query.filter_by(id=env.discord_id).first()
                        if discord:
                            tab.ele("@name=email").input(discord.name)
                            tab.ele("@name=password").input(aesCbcPbkdf2DecryptFromBase64(discord.pwd))
                            tab.ele("@type=submit").click()
                            fa2 = aesCbcPbkdf2DecryptFromBase64(discord.fa2)
                            if "login" in tab.url and len(fa2) > 10:
                                res = requests.get(fa2)
                                if res.ok:
                                    code = res.json().get("data").get("otp")
                                    tab.ele("@autocomplete=one-time-code").input(code)
                                    tab.ele("@type=submit").click()
                        else:
                            updateAccountStatus(env.discord_id, 1, "没有导入DISCORD 的账号信息")
                            raise Exception(f"{env.name}: 没有导入DISCORD 账号信息")

                if "channels" in tab.url or ".com/app" in tab.url:
                    updateAccountStatus(env.discord_id, 2)
                    logger.info(f"{env.name}登录Discord成功！")
        except Exception as e:
            pass
        getDiscord(chrome, env)
        tab.wait.load_start(timeout=10)

        if tab.wait.ele_displayed('@class=skip', timeout=10, raise_err=False):
            print('点击skip弹幕')
            tab.ele('@class=skip').click()
        else:
            print('没有出现skip')

        if tab.wait.ele_displayed('@class=skip', timeout=10, raise_err=False):
            print('点击skip弹幕')
            tab.ele('@class=skip').click()

        if tab.wait.ele_displayed('skip', timeout=10, raise_err=False):
            print('点击skip弹幕')
            tab.ele('@class=skip').click()
        tab.refresh()

        if tab.wait.ele_displayed('@class=bottom', timeout=60, raise_err=False):
            print('点击签到')
            with RedisLock(f"{env.name}-okx",200,200):
                tab.wait.load_start(timeout=5)
                tab.ele('@class=bottom').click(by_js=None)
                tab.wait.load_start(timeout=6)
                chrome.wait(3, 6)
                try:
                    chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
                    chrome.wait(15, 20)
                except Exception as e:
                    pass
        # elif tab.ele('@class=bottom disable'):
        #     print('已经签到过不需要签到了')

        else:
            tab.refresh()
            if tab.s_ele('t:p@text():Loading your profile...'):
                chrome.wait(15, 30)
                if tab.s_ele('t:p@text():Loading your profile...'):
                    print('1')
                    tab.refresh()
            tab.wait.ele_displayed('@class=bottom', timeout=30, raise_err=False)
            print('点击签到')
            tab.wait.load_start(timeout=5)
            tab.ele('@class=bottom').click(by_js=None)
            tab.wait.load_start(timeout=6)
            chrome.wait(3, 6)
            chrome.get_tab(title="OKX Wallet").ele("@type=button", index=2).click()
            chrome.wait(15, 20)

        if tab.wait.ele_displayed('@class=skip', timeout=15, raise_err=False):
            print('点击skip弹幕')
            tab.ele('@class=skip').click()

    global Rewards_Balance, Ranking, Rewards, Rewards_Yesterday, wallet
    try:
        taskData = getTaskObject(env, name)
        print('开始收集数据')
        if tab.wait.ele_displayed('@class=number', timeout=10, raise_err=False):
            Rewards_Balance = tab.ele('@class=number', index=1).text
            print('总分:', Rewards_Balance)
            Ranking = tab.ele('@class=number', index=2).text
            print('排名:', Ranking)
            Rewards = tab.ele('@class=number', index=3).text
            print('签到总分:', Rewards)
            Rewards_Yesterday = tab.ele('@class=number', index=4).text
            print('昨天得分:', Rewards_Yesterday)
            wallet = tab.ele('@class=chain', index=2).text
            print('wallet:', wallet)

        print('开始上传数据')
        taskData.Rewards_Balance = Rewards_Balance
        taskData.Ranking = Ranking
        taskData.Rewards = Rewards
        taskData.Rewards_Yesterday = Rewards_Yesterday
        taskData.wallet = wallet
        updateTaskRecord(env.name, name, taskData, 1)


        current_time = time.strftime("%m-%d")
        file_path = r'C:\Users\Public\Documents\humanity_{}.xlsx'.format(current_time)

        # 打开已存在的 Excel 文件（arch.xlsx）
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = '总分'
            ws['C1'] = '排名'
            ws['D1'] = '签到总分'
            ws['E1'] = '昨天得分'
            ws['F1'] = 'wallet'
        except FileNotFoundError:
            # 如果文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = '总分'
            ws['C1'] = '排名'
            ws['D1'] = '签到总分'
            ws['E1'] = '昨天得分'
            ws['F1'] = 'wallet'
            wb.save(file_path)
        # 找到下一行位置（避免覆盖）
        next_row = ws.max_row + 1

        env_name_exists = False
        env_name = env.name
        for row in range(2, ws.max_row + 1):  # 从第二行开始遍历（跳过表头）
            if ws[f'A{row}'].value == env_name:
                # 如果找到相同的 env_name，更新该行的 xp 和 level
                ws[f'B{row}'] = Rewards_Balance
                ws[f'C{row}'] = Ranking
                ws[f'D{row}'] = Rewards
                ws[f'E{row}'] = Rewards_Yesterday
                ws[f'F{row}'] = wallet

                env_name_exists = True
                break
        if not env_name_exists:
            # 如果没有找到相同的 env_name，追加新行
            next_row = ws.max_row + 1
            ws[f'A{next_row}'] = env_name
            ws[f'B{next_row}'] = Rewards_Balance
            ws[f'C{next_row}'] = Ranking
            ws[f'D{next_row}'] = Rewards
            ws[f'E{next_row}'] = Rewards_Yesterday
            ws[f'F{next_row}'] = wallet
        # 保存文件（不会覆盖，直接追加）
        wb.save(file_path)
    except Exception as e:
        logger.error(e)

def Humanity(chrome, env):
    # with app.app_context():
        try:
            # chrome: ChromiumPage = OKXChrome(env)
            gethumanity(chrome, env)
            logger.info(f"{env.name}环境：任务执行完毕，关闭环境")
        except Exception as e:
            logger.error(f"{env.name} 执行：{e}")
            return ("失败", e)
        # finally:
            # quitChrome(env, chrome)