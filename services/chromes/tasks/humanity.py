import string

import openpyxl
import requests
from DrissionPage import ChromiumPage,ChromiumOptions
from loguru import logger
# 连接数据库
from flaskServer.config.connect import app
from flaskServer.mode.account import Account
#登录环境账号
from flaskServer.services.chromes.login import OKXChrome
from flaskServer.services.dto.account import updateAccountStatus
from flaskServer.services.dto.task_record import updateTaskRecord, getTaskObject
import time
import random
from flaskServer.utils.chrome import quitChrome, get_Custome_Tab
from flaskServer.utils.crypt import aesCbcPbkdf2DecryptFromBase64

#humanity protocol
name = 'humanity'
humanity_url = 'https://testnet.humanity.org/login?ref=sunyunlei'

wallet_js = '''
let button  = 
document.querySelector("body > div:nth-child(31) > div > div > div._9pm4ki5.ju367va.ju367v15.ju367v8r > div > div > div > div > div > div.iekbcc0.ju367v6p._1vwt0cg2.ju367v7a.ju367v7v > div.iekbcc0.ju367va.ju367v15.ju367v1n > div:nth-child(1) > button > div > div");
button.click();
'''
dis_js = '''
let button  = 
document.querySelector("body > div.MuiBox-root.mui-186aokm > div > div > div > div.MuiBox-root.mui-1i3hx1p > div > div > div.MuiBox-root.mui-17age7i > a:nth-child(2)");
button.click();
'''

def exe_okx(chrome,env):
    try:
        if chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2):
            chrome.wait(3, 4)
            chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2).click()
            chrome.wait(3, 4)
            try:
                chrome.get_tab(title="OKX Wallet").ele('@data-testid=okd-button', index=2).click()
            except Exception as e:
                print('不需要二次确认')
    except Exception as e:
        print(f'{env.name}取的ele不对或者不需要连接')

def LoginDiscord(chrome:ChromiumPage,env):
    updateAccountStatus(env.discord_id, 0, "重置了Discord登录状态")
    tab = chrome.new_tab(url="https://discord.com/app")
    if tab.s_ele("Please log in again"):
        tab.ele("@class=button_dd4f85 lookFilled_dd4f85 colorPrimary_dd4f85 sizeMedium_dd4f85 grow_dd4f85").click()
    if "login" in tab.url:
        logger.info(f"{env.name} 开始登录 Discord 账号")
        with app.app_context():
            discord:Account = Account.query.filter_by(id=env.discord_id).first()
            if discord:
                tab.ele("@name=email").input(discord.name)
                tab.ele("@name=password").input(aesCbcPbkdf2DecryptFromBase64(discord.pwd))
                tab.ele("@type=submit").click()
                fa2 = aesCbcPbkdf2DecryptFromBase64(discord.fa2)
                if "login" in tab.url and len(fa2) > 10:
                    res = requests.get(fa2)
                    if res.ok:
                        code = res.json().get("data").get("otp")
                        tab.ele("@autocomplete=one-time-code").input(code)
                        tab.ele("@type=submit").click()
            else:
                updateAccountStatus(env.discord_id, 1, "没有导入DISCORD 的账号信息")
                raise Exception(f"{env.name}: 没有导入DISCORD 账号信息")
    if "channels" in tab.url or ".com/app" in tab.url:
        updateAccountStatus(env.discord_id, 2)
        logger.info(f"{env.name}登录Discord成功！")
    return get_Custome_Tab(tab)

def generate_random_word(length=7):
    # 生成一个随机的字母单词
    letters = string.ascii_lowercase + string.digits  # 包含小写字母、大写字母和数字
    return ''.join(random.choice(letters) for _ in range(length))

def getDiscord(chrome,env):
    try:
        print('开始discord认证')
        if chrome.get_tab(title='Discord | Authorize access to your account'):
            print('Discord | Authorize access to your account进入')
            chrome.get_tab(title='Discord | Authorize access to your account').ele("@type=button", index=2).click()
            logger.info(f"{env.name}: 登录discord完成----------------------------------")
            time.sleep(10)
        elif chrome.get_tab(title='Discord | 授权访问您的账号'):
            print('Discord | 授权访问您的账号进入')
            chrome.get_tab(title='Discord | 授权访问您的账号').ele("@type=button", index=2).click()
            logger.info(f"{env.name}: 登录discord完成----------------------------------")
            time.sleep(10)
        elif chrome.get_tab(title='Discord'):
            print('Discord 进入')
            if chrome.get_tab(title='Discord').ele('@class=button_dd4f85 lookFilled_dd4f85 colorPrimary_dd4f85 sizeMedium_dd4f85 grow_dd4f85'):
                logger.info(f'{env.name}的discord需要重新登录')
                chrome.get_tab(title='Discord').ele('@class=button_dd4f85 lookFilled_dd4f85 colorPrimary_dd4f85 sizeMedium_dd4f85 grow_dd4f85').click()
                time.sleep(6)
                LoginDiscord(chrome, env)
                time.sleep(6)
                chrome.close_tabs()
            elif chrome.get_tab(title='Discord').ele('Welcome back!'):
                logger.info(f"{env.name}的Discord未登录，尝试重新登录")
                time.sleep(6)
                LoginDiscord(chrome, env)
                time.sleep(6)
                chrome.close_tabs()
            if chrome.get_tab(title='Discord').ele("@type=button", index=2):
                chrome.get_tab(title='Discord').ele("@type=button", index=2).click()
                logger.info(f"{env.name}: 登录discord完成----------------------------------")
                time.sleep(10)
        elif chrome.get_tab('Onboarding | Humanity Protocol'):
            print('Onboarding | Humanity Protocol进去的')
            chrome.get_tab(title='Onboarding | Humanity Protocol').ele("@type=button", index=2).click()
            logger.info(f"{env.name}: 登录discord完成----------------------------------")
        else:
            logger.info(f'{env.name}:还有其他的语言需要加判断')
    except Exception as e:
        logger.error(e)

def gethumanity(chrome,env):
    tab = chrome.new_tab(url=humanity_url)
    tab.set.window.max()
    time.sleep(2)
    try:
        if tab.s_ele('@class=bottom'):
            print('点击签到')
            tab.ele('@class=bottom').click()
            time.sleep(5)
            try:
                exe_okx(chrome, env)
            except Exception as e:
                print('不需要连接钱包')
        elif tab.ele('@class=bottom disable'):
            print('已经签到过不需要签到了')
            return
    except Exception as e:
        logger.info(e)
    try:
        if tab.s_ele('Get Started'):
            try:
                tab.run_js(dis_js)
            except Exception as e:
                logger.info("之前登录过不需要再登录")
            time.sleep(2)
            #disocrd 授权过程
            try:
                logger.info(f'{env.name}:开始判断登录discord情况')
                chrome.wait(10, 20)
                getDiscord(chrome, env)
            except Exception as e:
                logger.error(e)
            time.sleep(10)
            try:
                if tab.s_ele('Choose a name for your Human ID'):
                    tab.ele('@class=MuiInputBase-input mui-1qvwndf').input(generate_random_word())
                    time.sleep(5)
                    tab.ele('@class=MuiBox-root mui-171onha', index=2).click()
                else:
                    time.sleep(10)
                    tab.ele('@class=MuiInputBase-input mui-1qvwndf').input(generate_random_word())
                    time.sleep(2)
                    tab.ele('@class=MuiBox-root mui-171onha', index=2).click()

                time.sleep(5)
                if tab.s_ele('Complete your profile'):
                    tab.ele('@class=MuiInputBase-input mui-1qvwndf', index=1).input(generate_random_word())
                    time.sleep(2)
                    tab.ele('@class=MuiInputBase-input mui-1qvwndf', index=2).input(generate_random_word())
                    time.sleep(2)
                    tab.ele('@class=MuiBox-root mui-171onha', index=2).click()
                else:
                    time.sleep(10)
                    tab.ele('@class=MuiInputBase-input mui-1qvwndf', index=1).input(generate_random_word())
                    time.sleep(5)
                    tab.ele('@class=MuiInputBase-input mui-1qvwndf', index=2).input(generate_random_word())
                    time.sleep(2)
                    tab.ele('@class=MuiBox-root mui-171onha', index=2).click()
            except Exception as e:
                logger.info(e)
        chrome.wait(10)
        try:
            if tab.s_ele('@class=skip'):
                print('点击skip弹幕')
                tab.ele('@class=skip').click()
            else:
                time.sleep(10)
                tab.ele('@class=skip').click()
        except Exception as e:
            print('没有出现skip')
        time.sleep(2)
        try:
            if tab.s_ele('@class=bottom'):
                print('点击签到')
                tab.ele('@class=bottom').click()
                time.sleep(5)
                try:
                    exe_okx(chrome, env)
                except Exception as e:
                    print('不需要连接钱包')
            elif tab.ele('@class=bottom disable'):
                print('已经签到过不需要签到了')
        except Exception as e:
            logger.info(e)
    except Exception as e:
        logger.error(e)

def getCount(chrome, env):
    try:
        taskData = getTaskObject(env, name)
        tab = chrome.new_tab(url=humanity_url)
        chrome.wait(10,20)
        print('开始收集数据')
        Rewards_Balance = tab.ele('@class=number', index=1).text
        print('总分:', Rewards_Balance)
        Ranking = tab.ele('@class=number', index=2).text
        print('排名:', Ranking)
        Rewards = tab.ele('@class=number', index=3).text
        print('签到总分:', Rewards)
        Rewards_Yesterday = tab.ele('@class=number', index=4).text
        print('昨天得分:', Rewards_Yesterday)
        wallet = tab.ele('@class=chain', index=2).text
        print('wallet:', wallet)

        print('开始上传数据')
        taskData.Rewards_Balance = Rewards_Balance
        taskData.Ranking = Ranking
        taskData.Rewards = Rewards
        taskData.Rewards_Yesterday = Rewards_Yesterday
        taskData.wallet = wallet
        updateTaskRecord(env.name, name, taskData, 1)


        current_time = time.strftime("%m-%d")
        file_path = r'C:\Users\Public\Documents\humanity_{}.xlsx'.format(current_time)

        # 打开已存在的 Excel 文件（arch.xlsx）
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = '总分'
            ws['C1'] = '排名'
            ws['D1'] = '签到总分'
            ws['E1'] = '昨天得分'
            ws['F1'] = 'wallet'
        except FileNotFoundError:
            # 如果文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            # 设置表头
            ws['A1'] = '环境编号'
            ws['B1'] = '总分'
            ws['C1'] = '排名'
            ws['D1'] = '签到总分'
            ws['E1'] = '昨天得分'
            ws['F1'] = 'wallet'
            wb.save(file_path)
        # 找到下一行位置（避免覆盖）
        next_row = ws.max_row + 1

        env_name_exists = False
        env_name = env.name
        for row in range(2, ws.max_row + 1):  # 从第二行开始遍历（跳过表头）
            if ws[f'A{row}'].value == env_name:
                # 如果找到相同的 env_name，更新该行的 xp 和 level
                ws[f'B{row}'] = Rewards_Balance
                ws[f'C{row}'] = Ranking
                ws[f'D{row}'] = Rewards
                ws[f'E{row}'] = Rewards_Yesterday
                ws[f'F{row}'] = wallet

                env_name_exists = True
                break
        if not env_name_exists:
            # 如果没有找到相同的 env_name，追加新行
            next_row = ws.max_row + 1
            ws[f'A{next_row}'] = env_name
            ws[f'B{next_row}'] = Rewards_Balance
            ws[f'C{next_row}'] = Ranking
            ws[f'D{next_row}'] = Rewards
            ws[f'E{next_row}'] = Rewards_Yesterday
            ws[f'F{next_row}'] = wallet

        # 保存文件（不会覆盖，直接追加）
        wb.save(file_path)
    except Exception as e:
        logger.error(e)




def Humanity(env):
    with app.app_context():
        try:
            chrome: ChromiumPage = OKXChrome(env)
            gethumanity(chrome, env)
            getCount(chrome, env)
            logger.info(f"{env.name}环境：任务执行完毕，关闭环境")
        except Exception as e:
            logger.error(f"{env.name} 执行：{e}")
            return ("失败", e)
        finally:
            quitChrome(env, chrome)